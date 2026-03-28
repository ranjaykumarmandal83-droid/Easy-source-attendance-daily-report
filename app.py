from flask import Flask, render_template, render_template_string, request, redirect, url_for, session, jsonify, send_file
import sqlite3, os, json, smtplib, secrets, hashlib
import io
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, date, timedelta
from functools import wraps

app = Flask(__name__)
app.secret_key = 'easysource_hrms_2024_xK9#mP2@vL5qN8wR'
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE']   = False   # True only if HTTPS
app.config['PERMANENT_SESSION_LIFETIME'] = 86400 * 7  # 7 days

DB = 'database/hrms.db'

def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def rows(cursor_result):
    """Convert sqlite3.Row list → plain list of dicts (fixes JSON serialization)"""
    return [dict(r) for r in cursor_result] if cursor_result else []

def row(r):
    return dict(r) if r else None

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT DEFAULT 'viewer',
            name TEXT,
            phone TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id TEXT UNIQUE,
            name TEXT NOT NULL,
            phone TEXT,
            department TEXT,
            designation TEXT,
            email TEXT,
            join_date TEXT,
            status TEXT DEFAULT 'active',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id TEXT,
            date TEXT,
            status TEXT,
            remark TEXT,
            marked_by TEXT,
            marked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS roster (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id TEXT,
            week_start TEXT,
            mon TEXT DEFAULT 'W', tue TEXT DEFAULT 'W',
            wed TEXT DEFAULT 'W', thu TEXT DEFAULT 'W',
            fri TEXT DEFAULT 'W', sat TEXT DEFAULT 'WO',
            sun TEXT DEFAULT 'WO'
        );
        CREATE TABLE IF NOT EXISTS whatsapp_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_name TEXT,
            group_id TEXT,
            message TEXT,
            sent_by TEXT,
            scheduled_time TEXT,
            sent_at TIMESTAMP,
            status TEXT DEFAULT 'pending',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS whatsapp_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            api_key TEXT,
            instance_id TEXT,
            auto_send_enabled INTEGER DEFAULT 0,
            auto_send_time TEXT DEFAULT '09:00',
            auto_message TEXT,
            groups TEXT
        );
        CREATE TABLE IF NOT EXISTS wa_groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT,
            group_type TEXT DEFAULT 'group',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS wa_schedules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id INTEGER,
            message TEXT,
            schedule_time TEXT,
            repeat_daily INTEGER DEFAULT 0,
            active INTEGER DEFAULT 1,
            last_sent TEXT,
            created_by TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        INSERT OR IGNORE INTO users (username, password, role, name) VALUES ('admin', 'admin123', 'admin', 'Administrator');
        INSERT OR IGNORE INTO users (username, password, role, name) VALUES ('user1', 'user123', 'user', 'User One');
    ''')
    conn.commit()

    # Migration: add missing columns to existing DBs
    migrations = [
        "ALTER TABLE employees ADD COLUMN location TEXT DEFAULT ''",
        "ALTER TABLE employees ADD COLUMN status TEXT DEFAULT 'active'",
        "ALTER TABLE employees ADD COLUMN office_setting TEXT DEFAULT ''",
        "ALTER TABLE employees ADD COLUMN gender TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN email TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN emp_id TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN can_whatsapp INTEGER DEFAULT 0",
        "ALTER TABLE users ADD COLUMN phone TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN name TEXT DEFAULT ''",
        "ALTER TABLE whatsapp_messages ADD COLUMN scheduled_time TEXT",
        "ALTER TABLE whatsapp_messages ADD COLUMN sent_at TIMESTAMP",
    ]
    for sql in migrations:
        try:
            conn.execute(sql)
            conn.commit()
        except Exception:
            pass  # Column already exists

    # Migrate old office_setting names to new names in existing DB data
    office_renames = [
        ('Head Office',          'Head Office Delhi'),
        ('Branch Office',        'Branch Office Gurgaon'),
        ('Branch Office Gurgoan','Branch Office Gurgaon'),  # typo fix
        ('Branch Office Gurgoan ','Branch Office Gurgaon'), # with trailing space
        ('Client Site',          'Client Side'),
        ('Work From Home',       'Work From Home'),         # keep same
    ]
    for old_name, new_name in office_renames:
        try:
            conn.execute("UPDATE employees SET office_setting=? WHERE office_setting=?", (new_name, old_name))
            conn.commit()
        except Exception:
            pass

    # Roster Rules table - employee-wise default weekly shift rules
    conn.execute('''CREATE TABLE IF NOT EXISTS roster_rules (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        emp_id TEXT UNIQUE NOT NULL,
        mon TEXT DEFAULT 'W',
        tue TEXT DEFAULT 'W',
        wed TEXT DEFAULT 'W',
        thu TEXT DEFAULT 'W',
        fri TEXT DEFAULT 'W',
        sat TEXT DEFAULT 'WO',
        sun TEXT DEFAULT 'WO',
        rule_name TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        created_by TEXT DEFAULT '',
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    # Roster Templates table - reusable named shift patterns
    conn.execute('''CREATE TABLE IF NOT EXISTS roster_templates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        description TEXT DEFAULT '',
        color TEXT DEFAULT 'primary',
        mon TEXT DEFAULT 'W',
        tue TEXT DEFAULT 'W',
        wed TEXT DEFAULT 'W',
        thu TEXT DEFAULT 'W',
        fri TEXT DEFAULT 'W',
        sat TEXT DEFAULT 'WO',
        sun TEXT DEFAULT 'WO',
        created_by TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    # Seed default roster templates if not exist
    DEFAULT_TEMPLATES = [
        ('Daily - As & When Required', '', 'success',
         'W','W','W','W','W','W','W'),
        ('Tuesdays & Fridays', '', 'primary',
         'WO','W','WO','WO','W','WO','WO'),
        ('Tuesdays & Thursdays', '', 'info',
         'WO','W','WO','W','WO','WO','WO'),
        ('Mondays, Wednesdays & Fridays', '', 'warning',
         'W','WO','W','WO','W','WO','WO'),
        ('Mondays, Wednesdays & Thursdays', '', 'danger',
         'W','WO','W','W','WO','WO','WO'),
        ('Daily - Client Site', '', 'dark',
         'OD','OD','OD','OD','OD','WO','WO'),
        ('WFH', '', 'secondary',
         'WFH','WFH','WFH','WFH','WFH','WO','WO'),
    ]
    for t in DEFAULT_TEMPLATES:
        existing = conn.execute('SELECT id FROM roster_templates WHERE name=?', (t[0],)).fetchone()
        if not existing:
            conn.execute(
                'INSERT INTO roster_templates (name,description,color,mon,tue,wed,thu,fri,sat,sun,created_by) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                (t[0], t[1], t[2], t[3], t[4], t[5], t[6], t[7], t[8], t[9], 'system')
            )
    conn.commit()

    # Heads table
    conn.execute('''CREATE TABLE IF NOT EXISTS heads (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER UNIQUE,
        emp_id TEXT UNIQUE,
        name TEXT,
        department TEXT,
        created_by TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS head_employees (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        head_emp_id TEXT,
        emp_id TEXT,
        UNIQUE(head_emp_id, emp_id)
    )''')
    conn.commit()

    # Email settings table
    conn.execute('''CREATE TABLE IF NOT EXISTS email_settings (
        id INTEGER PRIMARY KEY,
        smtp_host TEXT DEFAULT 'smtp.gmail.com',
        smtp_port INTEGER DEFAULT 587,
        smtp_user TEXT DEFAULT '',
        smtp_pass TEXT DEFAULT '',
        from_name TEXT DEFAULT 'EasySource HRMS',
        enabled INTEGER DEFAULT 0
    )''')
    conn.execute("INSERT OR IGNORE INTO email_settings (id) VALUES (1)")
    conn.commit()

    # Password reset tokens table
    conn.execute('''CREATE TABLE IF NOT EXISTS password_reset (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        token TEXT UNIQUE,
        expires_at TIMESTAMP,
        used INTEGER DEFAULT 0
    )''')
    conn.commit()

    # Auto-create user accounts for all active employees (emp_id as username+password)
    emps = conn.execute('SELECT emp_id, name, email, phone FROM employees WHERE status="active"').fetchall()
    for emp in emps:
        try:
            conn.execute(
                'INSERT OR IGNORE INTO users (username, password, role, name, phone, email, emp_id, can_whatsapp) VALUES (?,?,?,?,?,?,?,0)',
                (str(emp['emp_id']), str(emp['emp_id']), 'user', emp['name'], emp['phone'] or '', emp['email'] or '', emp['emp_id'])
            )
        except Exception:
            pass
    conn.commit()
    conn.close()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': 'Session expired. Please login again.', 'redirect': '/login'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            # API request ke liye JSON return karo, page request ke liye redirect
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': 'Session expired. Please login again.', 'redirect': '/login'}), 401
            return redirect(url_for('login'))
        if session.get('role') not in ['admin', 'subadmin']:
            return jsonify({'error': 'Access denied'}), 403
        return f(*args, **kwargs)
    return decorated

# ===== AUTH =====
@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '').strip()
        conn = get_db()
        u = conn.execute(
            'SELECT * FROM users WHERE LOWER(TRIM(username))=? AND TRIM(password)=?',
            (username, password)
        ).fetchone()
        conn.close()
        if u:
            session.permanent   = True
            session['user_id']  = u['id']
            session['username'] = u['username']
            session['role']     = u['role']
            session['name']     = u['name']
            session['emp_id']   = u['emp_id'] or ''
            session['can_whatsapp'] = bool(u['can_whatsapp'])
            # User role → redirect to mobile/self page
            if u['role'] == 'user':
                return redirect(url_for('my_attendance'))
            return redirect(url_for('dashboard'))
        return render_template('login.html', error='Invalid credentials. Username/Password check karo.')
    return render_template('login.html')

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    msg = None
    if request.method == 'POST':
        email = request.form.get('email','').strip().lower()
        conn  = get_db()
        u = conn.execute('SELECT * FROM users WHERE LOWER(TRIM(email))=?', (email,)).fetchone()
        if u:
            token = secrets.token_urlsafe(32)
            expires = (datetime.now() + timedelta(hours=2)).strftime('%Y-%m-%d %H:%M:%S')
            conn.execute('INSERT INTO password_reset (user_id,token,expires_at) VALUES (?,?,?)',
                         (u['id'], token, expires))
            conn.commit()
            # Send reset email
            reset_url = f"http://localhost:5000/reset-password/{token}"
            html = f"""
            <div style="font-family:Arial,sans-serif;max-width:480px;margin:0 auto;padding:24px;background:#f9f9f9;border-radius:12px">
              <h2 style="color:#1a3a5c">🔐 Password Reset</h2>
              <p>Namaste <b>{u['name'] or u['username']}</b>,</p>
              <p>Neeche diye link se apna password reset karo (2 ghante valid hai):</p>
              <a href="{reset_url}" style="display:inline-block;padding:12px 24px;background:#00c896;color:#fff;border-radius:8px;text-decoration:none;font-weight:bold;margin:16px 0">🔗 Reset Password</a>
              <p style="color:#888;font-size:12px">Agar aapne yeh request nahi ki toh ignore karein.</p>
              <p style="color:#888;font-size:12px">— EasySource HRMS</p>
            </div>"""
            send_email(u['email'], '🔐 HRMS Password Reset', html)
            msg = ('success', f'Reset link {email} par bhej diya gaya hai.')
        else:
            msg = ('error', 'Yeh email registered nahi hai.')
        conn.close()
    return render_template('forgot_password.html', msg=msg)

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    conn  = get_db()
    rec   = conn.execute(
        'SELECT * FROM password_reset WHERE token=? AND used=0 AND expires_at > datetime("now")',
        (token,)
    ).fetchone()
    if not rec:
        conn.close()
        return render_template('forgot_password.html', msg=('error', 'Link expired ya invalid hai. Dobara try karo.'))
    error = None
    if request.method == 'POST':
        new_pw = request.form.get('new_password','').strip()
        if len(new_pw) < 4:
            error = 'Password kam se kam 4 characters ka hona chahiye'
        else:
            conn.execute('UPDATE users SET password=? WHERE id=?', (new_pw, rec['user_id']))
            conn.execute('UPDATE password_reset SET used=1 WHERE id=?', (rec['id'],))
            conn.commit()
            conn.close()
            return render_template('forgot_password.html',
                                   msg=('success', '✅ Password reset ho gaya! Ab login karo.'))
    conn.close()
    return render_template('reset_password.html', token=token, error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ===== DASHBOARD =====
@app.route('/dashboard')
@login_required
def dashboard():
    if session.get('role') == 'user':
        return redirect(url_for('my_attendance'))
    return render_template('dashboard.html')

@app.route('/api/dashboard/stats', methods=['POST'])
@login_required
def api_dashboard_stats():
    data = request.json or {}
    today = date.today().isoformat()
    from_date = data.get('from_date', today)
    to_date   = data.get('to_date',   today)
    office    = data.get('office_setting')
    conn = get_db()

    # Build employee filter
    emp_where = 'status="active"'
    emp_params = []
    if office:
        emp_where += ' AND office_setting=?'
        emp_params.append(office)

    total_emp = conn.execute(f'SELECT COUNT(*) FROM employees WHERE {emp_where}', emp_params).fetchone()[0]

    # Attendance counts for date range
    def att_count(status):
        q = f'SELECT COUNT(DISTINCT a.id) FROM attendance a JOIN employees e ON a.emp_id=e.emp_id WHERE a.date BETWEEN ? AND ? AND a.status=? AND e.{emp_where}'
        return conn.execute(q, [from_date, to_date, status] + emp_params).fetchone()[0]

    present = att_count('P')
    absent  = att_count('A')
    od      = att_count('OD')
    leave   = att_count('L')

    # Office-wise stats (always for today)
    offices = ['Head Office Delhi','Branch Office Gurgaon','Client Side','Work From Home','Field']
    # Also normalize any remaining typos in DB at query time
    try:
        for typo, correct in [('Branch Office Gurgoan','Branch Office Gurgaon'),('Head Office','Head Office Delhi'),('Client Site','Client Side'),('Branch Office','Branch Office Gurgaon')]:
            conn.execute('UPDATE employees SET office_setting=? WHERE office_setting=?',(correct,typo))
        conn.commit()
    except: pass
    office_stats = []
    for o in offices:
        o_total = conn.execute('SELECT COUNT(*) FROM employees WHERE status="active" AND office_setting=?', (o,)).fetchone()[0]
        # Show all offices even if 0 employees
        def o_count(st, _office=o):   # _office captures correct value per iteration
            return conn.execute(
                'SELECT COUNT(*) FROM attendance a JOIN employees e ON a.emp_id=e.emp_id WHERE a.date BETWEEN ? AND ? AND a.status=? AND e.office_setting=?',
                (from_date, to_date, st, _office)).fetchone()[0]
        office_stats.append({
            'name': o, 'total': o_total,
            'present': o_count('P'), 'absent': o_count('A'),
            'od': o_count('OD'), 'leave': o_count('L')
        })

    conn.close()
    return jsonify({
        'total_emp': total_emp, 'present': present, 'absent': absent,
        'od': od, 'leave': leave, 'office_stats': office_stats
    })

# ===== EMPLOYEES =====
@app.route('/api/employees/sample-format')
@login_required  
def api_employees_sample_format():
    """CSV download - works on HTTP without any browser block"""
    import csv
    
    headers = ['emp_id','name','designation','department','location',
               'office_setting','phone','email','join_date','status','gender']
    
    sample_data = [
        ['5816','Sumit Kumar Mallick','DGM - Payroll & Benefits','Payroll','Gurugram Sector 44','Head Office Delhi','9876543210','sumit@easysource.in','01-Jan-2015','active','M'],
        ['0701','Prithvi Raj Chauhan','GM - Payroll & Benefits','Payroll','Gurugram Sector 18','Head Office Delhi','9871234567','prithvi@easysource.in','21-Sep-2009','active','M'],
        ['3534','Ranjay Kumar','Senior Manager - MIS','MIS','Gurugram','Head Office Delhi','9812345678','ranjay@easysource.in','01-Jul-2013','active','M'],
        ['3462','Dheeraj Mani Tripathi','DGM - Finance & Accounts','Finance & Accounts','Gurugram','Head Office Delhi','9873699560','dheeraj@easysource.in','17-Jun-2013','active','M'],
        ['52108','Abhishek Panday','Executive - Client Services','Operations','Delhi','Client Side','9878901234','abhishek@easysource.in','01-Dec-2023','active','M'],
    ]
    
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(sample_data)
    
    output = io.BytesIO()
    output.write(buf.getvalue().encode('utf-8-sig'))  # utf-8-sig for Excel compatibility
    output.seek(0)
    
    from flask import send_file
    return send_file(
        output,
        as_attachment=True,
        download_name='HRMS_Employee_Import_Format.csv',
        mimetype='text/csv'
    )

@app.route('/employees')
@login_required
def employees():
    if session.get('role') not in ['admin', 'subadmin']:
        return redirect(url_for('my_attendance'))
    conn = get_db()
    emps = rows(conn.execute('SELECT * FROM employees WHERE status="active" ORDER BY name').fetchall())
    # Pass set of emp_ids who are already heads (to show badge on card)
    head_ids = set(r[0] for r in conn.execute('SELECT emp_id FROM heads').fetchall())
    conn.close()
    return render_template('employees.html', employees=emps, head_ids=head_ids)

@app.route('/api/employees/counts')
@login_required
def api_emp_counts():
    conn = get_db()
    total    = conn.execute('SELECT COUNT(*) FROM employees').fetchone()[0]
    active   = conn.execute('SELECT COUNT(*) FROM employees WHERE status="active"').fetchone()[0]
    inactive = conn.execute('SELECT COUNT(*) FROM employees WHERE status="inactive"').fetchone()[0]
    conn.close()
    return jsonify({'total': total, 'active': active, 'inactive': inactive})

@app.route('/api/employees', methods=['GET'])
@login_required
def api_get_employees():
    conn = get_db()
    emps = rows(conn.execute('SELECT * FROM employees WHERE status="active" ORDER BY name').fetchall())
    conn.close()
    return jsonify(emps)

@app.route('/api/employees/add', methods=['POST'])
@admin_required
def api_add_employee():
    data = request.json
    conn = get_db()
    try:
        conn.execute(
            'INSERT INTO employees (emp_id, name, phone, department, designation, email, join_date, location, office_setting) VALUES (?,?,?,?,?,?,?,?,?)',
            (data['emp_id'], data['name'], data.get('phone',''), data.get('department',''),
             data.get('designation',''), data.get('email',''), data.get('join_date',''),
             data.get('location',''), data.get('office_setting',''))
        )
        # Auto-create login for this employee (emp_id = username = default password)
        try:
            conn.execute(
                'INSERT OR IGNORE INTO users (username, password, role, name, phone, email, emp_id, can_whatsapp) VALUES (?,?,?,?,?,?,?,0)',
                (data['emp_id'], data['emp_id'], 'user', data['name'],
                 data.get('phone',''), data.get('email',''), data['emp_id'])
            )
        except Exception:
            pass
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/employees/edit', methods=['POST'])
@admin_required
def api_edit_employee():
    data = request.json
    conn = get_db()
    conn.execute(
        'UPDATE employees SET name=?, phone=?, department=?, designation=?, email=?, join_date=?, location=?, office_setting=? WHERE id=?',
        (data['name'], data.get('phone',''), data.get('department',''),
         data.get('designation',''), data.get('email',''), data.get('join_date',''),
         data.get('location',''), data.get('office_setting',''), data['id'])
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/employees/delete', methods=['POST'])
@admin_required
def api_delete_employee():
    data = request.json
    conn = get_db()
    conn.execute('UPDATE employees SET status="inactive" WHERE id=?', (data['id'],))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/employees/bulk', methods=['POST'])
@admin_required
def api_bulk_employees():
    payload = request.json
    # Accept both direct list or {employees: [...]}
    if isinstance(payload, list):
        data = payload
    elif isinstance(payload, dict):
        data = payload.get('employees', [])
    else:
        data = []

    if not data:
        return jsonify({'error': 'Koi data nahi mila. Sahi Excel format use karo.', 'added': 0}), 400

    conn = get_db()
    added = 0
    skipped = 0
    errors = []

    def get_field(row, *keys):
        for k in keys:
            if k in row and str(row[k]).strip() not in ('', 'None', 'nan', 'NaN'):
                return str(row[k]).strip()
        return ''

    SKIP_VALUES = {'emp_id','Emp ID','EMP_ID','id','ID','name','Full Name','NAME',
                   'Employee Name','Required','REQUIRED','⭐ REQD','emp_id *','name *'}

    for i, r in enumerate(data):
        try:
            raw_id   = get_field(r, 'emp_id', 'Emp ID', 'EMP_ID', 'Emp. Code', 'EmpID', 'id', 'ID', 'Emp_ID')
            raw_name = get_field(r, 'name', 'Full Name', 'NAME', 'Employee Name', 'emp_name')

            if not raw_name or raw_name in SKIP_VALUES: skipped += 1; continue
            if not raw_id   or raw_id   in SKIP_VALUES: skipped += 1; continue

            phone    = get_field(r, 'phone', 'Phone', 'PHONE', 'Mobile', 'Phone Number', 'mobile')
            dept     = get_field(r, 'department', 'Department', 'DEPARTMENT', 'Dept', 'dept')
            desig    = get_field(r, 'designation', 'Designation', 'DESIGNATION', 'Desig', 'designaton')
            email    = get_field(r, 'email', 'Email', 'EMAIL', 'Email Address', 'email_id')
            jdate    = get_field(r, 'join_date', 'Join Date', 'JOIN_DATE', 'Joining Date', 'joining_date', 'date_of_joining')
            status   = get_field(r, 'status', 'Status', 'STATUS') or 'active'
            location = get_field(r, 'location', 'Location', 'LOCATION', 'City', 'city', 'Office Location') or ''
            office_s = get_field(r, 'office_setting', 'Office Setting', 'OFFICE_SETTING', 'Office', 'office_type', 'office') or ''
            gender   = get_field(r, 'gender', 'Gender', 'GENDER', 'sex', 'Sex') or ''

            # Try INSERT, if emp_id exists then UPDATE
            existing = conn.execute('SELECT id FROM employees WHERE emp_id=?', (raw_id,)).fetchone()
            if existing:
                conn.execute(
                    'UPDATE employees SET name=?,phone=?,department=?,designation=?,email=?,join_date=?,status=?,location=?,office_setting=?,gender=? WHERE emp_id=?',
                    (raw_name, phone, dept, desig, email, jdate, status, location, office_s, gender, raw_id)
                )
            else:
                conn.execute(
                    'INSERT INTO employees (emp_id,name,phone,department,designation,email,join_date,status,location,office_setting,gender) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                    (raw_id, raw_name, phone, dept, desig, email, jdate, status, location, office_s, gender)
                )
            # Auto-create login account (emp_id = username = default password)
            try:
                conn.execute(
                    'INSERT OR IGNORE INTO users (username,password,role,name,phone,email,emp_id,can_whatsapp) VALUES (?,?,?,?,?,?,?,0)',
                    (raw_id, raw_id, 'user', raw_name, phone, email, raw_id)
                )
            except Exception:
                pass
            added += 1
        except Exception as e:
            errors.append(f'Row {i+2}: {str(e)}')

    conn.commit()
    conn.close()
    return jsonify({
        'success': True,
        'added': added,
        'skipped': skipped,
        'errors': errors[:5]
    })

# ===== ATTENDANCE =====
@app.route('/attendance')
@login_required
def attendance():
    if session.get('role') not in ['admin', 'subadmin']:
        return redirect(url_for('my_attendance'))
    conn = get_db()
    emps = rows(conn.execute('SELECT emp_id, name, department, designation, office_setting FROM employees WHERE status="active" ORDER BY name').fetchall())
    conn.close()
    today = date.today().isoformat()
    return render_template('attendance.html', employees=emps, now=today)

@app.route('/api/attendance/get', methods=['POST'])
@login_required
def api_get_attendance():
    data = request.json
    conn = get_db()
    records = rows(conn.execute('SELECT * FROM attendance WHERE date=?', (data['date'],)).fetchall())
    conn.close()
    return jsonify(records)

@app.route('/api/attendance/mark', methods=['POST'])
@login_required
def api_mark_attendance():
    data = request.json
    conn = get_db()
    existing = conn.execute('SELECT id FROM attendance WHERE emp_id=? AND date=?',
                            (data['emp_id'], data['date'])).fetchone()
    if existing:
        conn.execute(
            'UPDATE attendance SET status=?, remark=?, marked_by=?, marked_at=CURRENT_TIMESTAMP WHERE id=?',
            (data['status'], data.get('remark',''), session['username'], existing['id'])
        )
    else:
        conn.execute(
            'INSERT INTO attendance (emp_id, date, status, remark, marked_by) VALUES (?,?,?,?,?)',
            (data['emp_id'], data['date'], data['status'], data.get('remark',''), session['username'])
        )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/attendance/bulk', methods=['POST'])
@login_required
def api_bulk_attendance():
    data = request.json
    conn = get_db()
    for rec in data['records']:
        existing = conn.execute('SELECT id FROM attendance WHERE emp_id=? AND date=?',
                                (rec['emp_id'], data['date'])).fetchone()
        if existing:
            conn.execute(
                'UPDATE attendance SET status=?, remark=?, marked_by=? WHERE id=?',
                (rec['status'], rec.get('remark',''), session['username'], existing['id'])
            )
        else:
            conn.execute(
                'INSERT INTO attendance (emp_id, date, status, remark, marked_by) VALUES (?,?,?,?,?)',
                (rec['emp_id'], data['date'], rec['status'], rec.get('remark',''), session['username'])
            )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ===== ROSTER =====
def seed_default_templates(conn):
    count = conn.execute('SELECT COUNT(*) FROM roster_templates').fetchone()[0]
    if count == 0:
        DEFAULT_TEMPLATES = [
            ('Daily - As & When Required', '', 'success','W','W','W','W','W','W','W'),
            ('Tuesdays & Fridays', '', 'primary','WO','W','WO','WO','W','WO','WO'),
            ('Tuesdays & Thursdays', '', 'info','WO','W','WO','W','WO','WO','WO'),
            ('Mondays, Wednesdays & Fridays', '', 'warning','W','WO','W','WO','W','WO','WO'),
            ('Mondays, Wednesdays & Thursdays', '', 'danger','W','WO','W','W','WO','WO','WO'),
            ('Daily - Client Site', '', 'dark','OD','OD','OD','OD','OD','WO','WO'),
            ('WFH', '', 'secondary','WFH','WFH','WFH','WFH','WFH','WO','WO'),
        ]
        for t in DEFAULT_TEMPLATES:
            conn.execute(
                'INSERT OR IGNORE INTO roster_templates (name,description,color,mon,tue,wed,thu,fri,sat,sun,created_by) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                (t[0],t[1],t[2],t[3],t[4],t[5],t[6],t[7],t[8],t[9],'system')
            )
        conn.commit()

@app.route('/roster')
@login_required
def roster():
    role = session.get('role')
    if role not in ['admin', 'subadmin', 'head']:
        return redirect(url_for('my_attendance'))
    conn = get_db()
    seed_default_templates(conn)
    if role == 'head':
        # Head sees only their linked employees
        head_emp_id = session.get('emp_id', '')
        linked_ids = [r[0] for r in conn.execute(
            'SELECT emp_id FROM head_employees WHERE head_emp_id=?', (head_emp_id,)
        ).fetchall()]
        if linked_ids:
            placeholders = ','.join(['?' for _ in linked_ids])
            emps = rows(conn.execute(
                f'SELECT id,emp_id,name,department,designation,location,office_setting FROM employees WHERE emp_id IN ({placeholders}) AND status="active" ORDER BY name',
                linked_ids
            ).fetchall())
        else:
            emps = []
    else:
        emps = rows(conn.execute('SELECT id, emp_id, name, department, designation, location, office_setting FROM employees WHERE status="active" ORDER BY name').fetchall())
    conn.close()
    return render_template('roster.html', employees=emps)

@app.route('/api/attendance/emp-history', methods=['POST'])
@login_required
def api_emp_history():
    data = request.json
    conn = get_db()
    records = rows(conn.execute(
        'SELECT date, status, remark, marked_by FROM attendance WHERE emp_id=? AND date BETWEEN ? AND ? ORDER BY date DESC',
        (data['emp_id'], data['from_date'], data['to_date'])
    ).fetchall())
    conn.close()
    return jsonify(records)

@app.route('/api/roster/get', methods=['POST'])
@login_required
def api_get_roster():
    data = request.json
    conn = get_db()
    week_start = data.get('week_start', '')
    week_end   = data.get('week_end', week_start)
    if week_end and week_end != week_start:
        records = rows(conn.execute(
            'SELECT * FROM roster WHERE week_start BETWEEN ? AND ? ORDER BY week_start',
            (week_start, week_end)).fetchall())
    else:
        records = rows(conn.execute(
            'SELECT * FROM roster WHERE week_start=?', (week_start,)).fetchall())
    conn.close()
    return jsonify(records)



# ===== WHATSAPP =====
@app.route('/whatsapp')
@login_required
def whatsapp():
    role = session.get('role')
    if role not in ['admin', 'subadmin']:
        # Check can_whatsapp permission
        conn2 = get_db()
        u = conn2.execute('SELECT can_whatsapp FROM users WHERE id=?', (session['user_id'],)).fetchone()
        conn2.close()
        if not u or not u['can_whatsapp']:
            return render_template('error.html', message='WhatsApp access ki permission nahi hai. Admin se contact karein.') if os.path.exists('templates/error.html') else ('⛔ Access Denied', 403)
    conn = get_db()
    messages = rows(conn.execute('SELECT * FROM whatsapp_messages ORDER BY created_at DESC LIMIT 50').fetchall())
    settings = row(conn.execute('SELECT * FROM whatsapp_settings LIMIT 1').fetchone())
    conn.close()
    return render_template('whatsapp.html', messages=messages, settings=settings)

@app.route('/api/whatsapp/send', methods=['POST'])
@login_required
def api_send_whatsapp():
    data = request.json
    conn = get_db()
    conn.execute(
        'INSERT INTO whatsapp_messages (group_name, group_id, message, sent_by, status) VALUES (?,?,?,?,?)',
        (data.get('group_name',''), data.get('group_id',''), data['message'], session['username'], 'sent')
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/whatsapp/schedule', methods=['POST'])
@login_required
def api_schedule_whatsapp():
    if session.get('role') not in ['admin', 'subadmin']:
        return jsonify({'error': 'Not allowed'}), 403
    data = request.json
    conn = get_db()
    conn.execute(
        'INSERT INTO whatsapp_messages (group_name, group_id, message, sent_by, scheduled_time, status) VALUES (?,?,?,?,?,?)',
        (data.get('group_name',''), data.get('group_id',''), data['message'],
         session['username'], data.get('scheduled_time'), 'scheduled')
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/whatsapp/settings', methods=['POST'])
@admin_required
def api_whatsapp_settings():
    data = request.json
    conn = get_db()
    existing = conn.execute('SELECT id FROM whatsapp_settings LIMIT 1').fetchone()
    groups_json = json.dumps(data.get('groups', []))
    if existing:
        conn.execute(
            'UPDATE whatsapp_settings SET api_key=?,instance_id=?,auto_send_enabled=?,auto_send_time=?,auto_message=?,groups=? WHERE id=?',
            (data.get('api_key',''), data.get('instance_id',''), data.get('auto_send_enabled',0),
             data.get('auto_send_time','09:00'), data.get('auto_message',''), groups_json, existing['id'])
        )
    else:
        conn.execute(
            'INSERT INTO whatsapp_settings (api_key,instance_id,auto_send_enabled,auto_send_time,auto_message,groups) VALUES (?,?,?,?,?,?)',
            (data.get('api_key',''), data.get('instance_id',''), data.get('auto_send_enabled',0),
             data.get('auto_send_time','09:00'), data.get('auto_message',''), groups_json)
        )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ===== WA GROUP MANAGEMENT =====
@app.route('/api/wa/groups', methods=['GET'])
@login_required
def api_wa_groups_get():
    conn = get_db()
    groups = rows(conn.execute('SELECT * FROM wa_groups ORDER BY name').fetchall())
    conn.close()
    return jsonify(groups)

@app.route('/api/wa/groups/add', methods=['POST'])
@login_required
def api_wa_groups_add():
    data = request.json
    name = data.get('name','').strip()
    phone = data.get('phone','').strip()
    gtype = data.get('group_type', 'group')
    if not name:
        return jsonify({'success': False, 'error': 'Name required'})
    conn = get_db()
    conn.execute('INSERT INTO wa_groups (name, phone, group_type) VALUES (?,?,?)', (name, phone, gtype))
    conn.commit()
    gid = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    conn.close()
    return jsonify({'success': True, 'id': gid})

@app.route('/api/wa/groups/delete', methods=['POST'])
@login_required
def api_wa_groups_delete():
    data = request.json
    conn = get_db()
    conn.execute('DELETE FROM wa_groups WHERE id=?', (data['id'],))
    conn.execute('DELETE FROM wa_schedules WHERE group_id=?', (data['id'],))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ===== WA SEND VIA WHATSAPP WEB (wa.me link) =====
@app.route('/api/wa/send_web', methods=['POST'])
@login_required
def api_wa_send_web():
    """Returns wa.me link OR pywhatkit command for the group"""
    data = request.json
    group_id = data.get('group_id')
    message  = data.get('message','').strip()
    conn = get_db()
    grp = row(conn.execute('SELECT * FROM wa_groups WHERE id=?', (group_id,)).fetchone())
    if not grp:
        conn.close()
        return jsonify({'success': False, 'error': 'Group not found'})
    # Save to history
    conn.execute(
        'INSERT INTO whatsapp_messages (group_name, group_id, message, sent_by, status) VALUES (?,?,?,?,?)',
        (grp['name'], str(group_id), message, session['username'], 'sent')
    )
    conn.commit()
    conn.close()
    import urllib.parse
    encoded = urllib.parse.quote(message)
    phone = (grp.get('phone') or '').strip().replace('+','').replace(' ','').replace('-','')
    if phone:
        wa_link = f"https://wa.me/{phone}?text={encoded}"
    else:
        wa_link = f"https://web.whatsapp.com/send?text={encoded}"
    return jsonify({'success': True, 'wa_link': wa_link, 'group_name': grp['name']})

# ===== WA SCHEDULER =====
@app.route('/api/wa/schedules', methods=['GET'])
@login_required
def api_wa_schedules_get():
    conn = get_db()
    scheds = rows(conn.execute('''
        SELECT s.*, g.name as group_name
        FROM wa_schedules s
        LEFT JOIN wa_groups g ON s.group_id = g.id
        ORDER BY s.schedule_time
    ''').fetchall())
    conn.close()
    return jsonify(scheds)

@app.route('/api/wa/schedules/add', methods=['POST'])
@login_required
def api_wa_schedules_add():
    data = request.json
    conn = get_db()
    conn.execute(
        'INSERT INTO wa_schedules (group_id, message, schedule_time, repeat_daily, active, created_by) VALUES (?,?,?,?,1,?)',
        (data['group_id'], data['message'], data['schedule_time'], 1 if data.get('repeat_daily') else 0, session['username'])
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/wa/schedules/toggle', methods=['POST'])
@login_required
def api_wa_schedules_toggle():
    data = request.json
    conn = get_db()
    conn.execute('UPDATE wa_schedules SET active=? WHERE id=?', (data['active'], data['id']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/wa/schedules/delete', methods=['POST'])
@login_required
def api_wa_schedules_delete():
    data = request.json
    conn = get_db()
    conn.execute('DELETE FROM wa_schedules WHERE id=?', (data['id'],))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/wa/check_schedules', methods=['GET'])
@login_required
def api_wa_check_schedules():
    """Check if any schedule should fire now — called by frontend every minute"""
    now = datetime.now()
    current_time = now.strftime('%H:%M')
    current_date = now.strftime('%Y-%m-%d')
    conn = get_db()
    due = rows(conn.execute('''
        SELECT s.*, g.name as group_name, g.phone as group_phone
        FROM wa_schedules s
        LEFT JOIN wa_groups g ON s.group_id = g.id
        WHERE s.active=1 AND s.schedule_time=?
          AND (s.last_sent IS NULL OR s.last_sent != ?)
    ''', (current_time, current_date)).fetchall())
    # Mark as sent
    for s in due:
        conn.execute('UPDATE wa_schedules SET last_sent=? WHERE id=?', (current_date, s['id']))
        conn.execute(
            'INSERT INTO whatsapp_messages (group_name, group_id, message, sent_by, status) VALUES (?,?,?,?,?)',
            (s['group_name'], str(s['group_id']), s['message'], 'auto-scheduler', 'sent')
        )
    conn.commit()
    conn.close()
    import urllib.parse
    results = []
    for s in due:
        encoded = urllib.parse.quote(s['message'])
        phone = (s.get('group_phone') or '').strip().replace('+','').replace(' ','').replace('-','')
        if phone:
            wa_link = f"https://wa.me/{phone}?text={encoded}"
        else:
            wa_link = f"https://web.whatsapp.com/send?text={encoded}"
        results.append({'group_name': s['group_name'], 'wa_link': wa_link, 'message': s['message']})
    return jsonify({'due': results})

# ===== ATTENDANCE WHATSAPP TEMPLATE =====
@app.route('/api/whatsapp/attendance_template', methods=['POST'])
@login_required
def api_attendance_template():
    data = request.json
    date_str = data.get('date', datetime.now().strftime('%Y-%m-%d'))
    use_roster = data.get('use_roster', False)
    conn = get_db()

    office_setting = data.get('office_setting')
    office_filter = ' AND e.office_setting=?' if office_setting else ''
    params_base = [date_str] + ([office_setting] if office_setting else [])

    if use_roster:
        # Get day of week for roster shift
        from datetime import datetime as dt2
        d = dt2.strptime(date_str, '%Y-%m-%d')
        day_col = ['mon','tue','wed','thu','fri','sat','sun'][d.weekday()]
        # Get week start (Monday)
        week_start = (d - timedelta(days=d.weekday())).strftime('%Y-%m-%d')
        records = conn.execute(f'''
            SELECT e.emp_id, e.name, e.designation, e.department, e.location,
                   e.office_setting, a.status, a.remark,
                   r.{day_col} as roster_shift
            FROM employees e
            LEFT JOIN attendance a ON e.emp_id=a.emp_id AND a.date=?
            LEFT JOIN roster r ON e.emp_id=r.emp_id AND r.week_start=?
            WHERE e.status="active"{office_filter}
            ORDER BY e.name
        ''', [date_str, week_start] + ([office_setting] if office_setting else [])).fetchall()
    else:
        records = conn.execute(f'''
            SELECT e.emp_id, e.name, e.designation, e.department, e.location,
                   e.office_setting, a.status, a.remark, NULL as roster_shift
            FROM employees e
            LEFT JOIN attendance a ON e.emp_id=a.emp_id AND a.date=?
            WHERE e.status="active"{office_filter}
            ORDER BY e.name
        ''', params_base).fetchall()
    conn.close()

    # Format date nicely
    try:
        from datetime import datetime as dt
        d = dt.strptime(date_str, '%Y-%m-%d')
        display_date = d.strftime('%d-%b-%Y')
    except:
        display_date = date_str

    STATUS_EMOJI = {
        'P':   '✅',
        'A':   '❌',
        'OD':  '🚗',
        'WFH': '🏠',
        'HD':  '🌓',
        'L':   '🏖️',
        None:  '❓',
        '':    '❓',
    }
    STATUS_LABEL = {
        'P':   'Present',
        'A':   'Absent',
        'OD':  'OD',
        'WFH': 'WFH',
        'HD':  'Half Day',
        'L':   'Leave',
        None:  'Not Marked',
        '':    'Not Marked',
    }

    lines = []
    lines.append(f"📊 *ATTENDANCE REPORT*")
    lines.append(f"📅 Date: {display_date}")
    if office_setting:
        lines.append(f"🏢 Office: {office_setting}")
    lines.append("━━━━━━━━━━━━━━━━━━━━")
    lines.append("")

    counts = {'P': 0, 'A': 0, 'OD': 0, 'WFH': 0, 'HD': 0, 'L': 0, 'Not Marked': 0}

    number_emojis = ['1️⃣','2️⃣','3️⃣','4️⃣','5️⃣','6️⃣','7️⃣','8️⃣','9️⃣','🔟']

    for i, r in enumerate(records):
        emp_id       = r[0]
        name         = r[1]
        status       = r[6] if len(r) > 6 else ''
        status       = status or ''
        remark       = r[7] if len(r) > 7 else ''
        remark       = remark or ''
        roster_shift = r[8] if len(r) > 8 else None
        emoji   = STATUS_EMOJI.get(status, '❓')
        label   = STATUS_LABEL.get(status, status)
        num     = number_emojis[i] if i < len(number_emojis) else f"{i+1}."

        remark_txt = f" — _{remark}_" if remark else ""
        roster_txt = f" [{roster_shift}]" if roster_shift and roster_shift not in ['W',''] else ""
        lines.append(f"{num} {emp_id} | {name} | {emoji} {label}{roster_txt}{remark_txt}")

        if status == 'P':   counts['P'] += 1
        elif status == 'A': counts['A'] += 1
        elif status == 'OD': counts['OD'] += 1
        elif status == 'WFH': counts['WFH'] += 1
        elif status == 'HD': counts['HD'] += 1
        elif status == 'L':  counts['L'] += 1
        else: counts['Not Marked'] += 1

    lines.append("")
    lines.append("━━━━━━━━━━━━━━━━━━━━")

    summary_parts = []
    if counts['P']:   summary_parts.append(f"✅ Present: {counts['P']}")
    if counts['A']:   summary_parts.append(f"❌ Absent: {counts['A']}")
    if counts['OD']:  summary_parts.append(f"🚗 OD: {counts['OD']}")
    if counts['WFH']: summary_parts.append(f"🏠 WFH: {counts['WFH']}")
    if counts['HD']:  summary_parts.append(f"🌓 Half Day: {counts['HD']}")
    if counts['L']:   summary_parts.append(f"🏖️ Leave: {counts['L']}")
    if counts['Not Marked']: summary_parts.append(f"❓ Not Marked: {counts['Not Marked']}")

    lines.append(" | ".join(summary_parts))
    lines.append(f"*Total: {len(records)} Employees*")

    message = "\n".join(lines)
    return jsonify({'success': True, 'message': message, 'date': display_date})

# ===== REPORTS =====
@app.route('/reports')
@login_required
def reports():
    if session.get('role') not in ['admin', 'subadmin']:
        return redirect(url_for('my_attendance'))
    return render_template('reports.html')

@app.route('/api/reports/attendance', methods=['POST'])
@login_required
def api_attendance_report():
    data = request.json
    conn = get_db()
    query = '''SELECT e.emp_id, e.name, e.department, e.office_setting, a.date, a.status, a.remark
               FROM employees e LEFT JOIN attendance a ON e.emp_id=a.emp_id
               WHERE a.date BETWEEN ? AND ?'''
    params = [data['from_date'], data['to_date']]
    if data.get('department'):
        query += ' AND e.department=?'
        params.append(data['department'])
    if data.get('office_setting'):
        query += ' AND e.office_setting=?'
        params.append(data['office_setting'])
    if data.get('emp_query'):
        query += ' AND (e.emp_id LIKE ? OR LOWER(e.name) LIKE ?)'
        q = '%' + data['emp_query'].lower() + '%'
        params.extend([q, q])
    query += ' ORDER BY e.name, a.date'
    records = rows(conn.execute(query, params).fetchall())
    conn.close()
    return jsonify(records)

# ===== HEADS =====
@app.route('/api/heads/download')
@admin_required
def api_heads_download():
    """Download Master Head List as CSV"""
    import csv
    conn = get_db()
    heads = rows(conn.execute('''
        SELECT h.id, h.emp_id, h.name, h.department, h.created_at, u.username
        FROM heads h LEFT JOIN users u ON h.user_id = u.id ORDER BY h.name
    ''').fetchall())
    for h in heads:
        linked = conn.execute('''
            SELECT e.emp_id, e.name, e.department
            FROM head_employees he JOIN employees e ON he.emp_id = e.emp_id
            WHERE he.head_emp_id = ?
        ''', (h['emp_id'],)).fetchall()
        h['linked'] = linked
    conn.close()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['Head Emp ID','Head Name','Department','Login Username',
                     'Created At','Total Linked','Linked Emp IDs','Linked Emp Names'])
    for h in heads:
        linked_ids   = ' | '.join([r[0] for r in h['linked']])
        linked_names = ' | '.join([r[1] for r in h['linked']])
        writer.writerow([
            h['emp_id'], h['name'], h['department'] or '',
            h['username'] or h['emp_id'],
            (h['created_at'] or '')[:10],
            len(h['linked']), linked_ids, linked_names
        ])

    output = io.BytesIO()
    output.write(buf.getvalue().encode('utf-8-sig'))
    output.seek(0)
    from flask import send_file
    return send_file(output, as_attachment=True,
                     download_name='HRMS_Master_Head_List.csv',
                     mimetype='text/csv')

@app.route('/api/heads/list', methods=['GET'])
@admin_required
def api_heads_list():
    conn = get_db()
    heads = rows(conn.execute('''
        SELECT h.id, h.emp_id, h.name, h.department, h.created_at,
               u.username, u.role
        FROM heads h
        LEFT JOIN users u ON h.user_id = u.id
        ORDER BY h.name
    ''').fetchall())
    for h in heads:
        # Get linked employees
        linked = rows(conn.execute('''
            SELECT e.emp_id, e.name, e.department, e.designation
            FROM head_employees he
            JOIN employees e ON he.emp_id = e.emp_id
            WHERE he.head_emp_id = ?
            ORDER BY e.name
        ''', (h['emp_id'],)).fetchall())
        h['linked_employees'] = linked
        h['linked_count'] = len(linked)
    conn.close()
    return jsonify(heads)

@app.route('/api/heads/create', methods=['POST'])
@admin_required
def api_create_head():
    data = request.json
    emp_id = data.get('emp_id', '').strip()
    if not emp_id:
        return jsonify({'error': 'Employee ID zaroori hai'}), 400
    conn = get_db()
    # Check duplicate
    existing = conn.execute('SELECT id FROM heads WHERE emp_id=?', (emp_id,)).fetchone()
    if existing:
        conn.close()
        return jsonify({'error': f'"{emp_id}" already Head hai. Duplicate nahi ban sakta!'}), 400
    # Get employee info
    emp = conn.execute('SELECT * FROM employees WHERE emp_id=?', (emp_id,)).fetchone()
    if not emp:
        conn.close()
        return jsonify({'error': 'Employee nahi mila'}), 404
    emp = dict(emp)
    # Ensure user account exists for this employee
    user = conn.execute('SELECT id FROM users WHERE emp_id=?', (emp_id,)).fetchone()
    if not user:
        conn.execute(
            'INSERT OR IGNORE INTO users (username, password, role, name, emp_id) VALUES (?,?,?,?,?)',
            (emp_id, emp_id, 'head', emp['name'], emp_id)
        )
        conn.commit()
        user = conn.execute('SELECT id FROM users WHERE username=?', (emp_id,)).fetchone()
    else:
        # Update role to head
        conn.execute('UPDATE users SET role="head" WHERE id=?', (user['id'],))
        conn.commit()
    user_id = user['id'] if user else None
    # Create head record
    conn.execute(
        'INSERT INTO heads (user_id, emp_id, name, department, created_by) VALUES (?,?,?,?,?)',
        (user_id, emp_id, emp['name'], emp.get('department',''), session['username'])
    )
    # Link employees if provided
    linked = data.get('linked_emp_ids', [])
    for eid in linked:
        try:
            conn.execute('INSERT OR IGNORE INTO head_employees (head_emp_id, emp_id) VALUES (?,?)', (emp_id, eid))
        except: pass
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': f'{emp["name"]} ko Head bana diya!'})

@app.route('/api/heads/delete', methods=['POST'])
@admin_required
def api_delete_head():
    data = request.json
    head_id = data.get('head_id')
    conn = get_db()
    head = conn.execute('SELECT * FROM heads WHERE id=?', (head_id,)).fetchone()
    if head:
        head = dict(head)
        # Revert role to subadmin
        conn.execute('UPDATE users SET role="subadmin" WHERE id=?', (head['user_id'],))
        conn.execute('DELETE FROM head_employees WHERE head_emp_id=?', (head['emp_id'],))
        conn.execute('DELETE FROM heads WHERE id=?', (head_id,))
        conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/heads/link-employees', methods=['POST'])
@admin_required
def api_head_link_employees():
    data = request.json
    head_emp_id = data.get('head_emp_id')
    emp_ids = data.get('emp_ids', [])
    conn = get_db()
    # Remove old links for this head
    conn.execute('DELETE FROM head_employees WHERE head_emp_id=?', (head_emp_id,))
    for eid in emp_ids:
        try:
            conn.execute('INSERT OR IGNORE INTO head_employees (head_emp_id, emp_id) VALUES (?,?)', (head_emp_id, eid))
        except: pass
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# Head can save roster for their linked employees
@app.route('/api/roster/save', methods=['POST'])
@login_required
def api_save_roster():
    if session.get('role') not in ['admin', 'subadmin', 'head']:
        return jsonify({'error': 'Permission denied'}), 403
    data = request.json
    conn = get_db()
    # Head can only save roster for their linked employees
    if session.get('role') == 'head':
        head_emp_id = session.get('emp_id')
        allowed = [r[0] for r in conn.execute(
            'SELECT emp_id FROM head_employees WHERE head_emp_id=?', (head_emp_id,)
        ).fetchall()]
        if data['emp_id'] not in allowed:
            conn.close()
            return jsonify({'error': 'Aap sirf apne linked employees ka roster save kar sakte ho'}), 403
    existing = conn.execute('SELECT id FROM roster WHERE emp_id=? AND week_start=?',
                            (data['emp_id'], data['week_start'])).fetchone()
    days = ['mon','tue','wed','thu','fri','sat','sun']
    if existing:
        vals = [data.get(d,'W') for d in days]
        conn.execute('UPDATE roster SET mon=?,tue=?,wed=?,thu=?,fri=?,sat=?,sun=? WHERE id=?',
                     vals + [existing['id']])
    else:
        conn.execute(
            'INSERT INTO roster (emp_id,week_start,mon,tue,wed,thu,fri,sat,sun) VALUES (?,?,?,?,?,?,?,?,?)',
            (data['emp_id'], data['week_start'],
             data.get('mon','W'), data.get('tue','W'), data.get('wed','W'),
             data.get('thu','W'), data.get('fri','W'), data.get('sat','WO'), data.get('sun','WO'))
        )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ===== SETTINGS / USERS =====
# ===== EMAIL HELPER =====
def send_email(to_email, subject, body_html):
    try:
        conn = get_db()
        cfg = conn.execute('SELECT * FROM email_settings WHERE id=1').fetchone()
        conn.close()
        if not cfg or not cfg['enabled'] or not cfg['smtp_user']:
            return False, 'Email not configured'
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f"{cfg['from_name']} <{cfg['smtp_user']}>"
        msg['To']      = to_email
        msg.attach(MIMEText(body_html, 'html'))
        port = int(cfg['smtp_port'])
        if port == 465:
            with smtplib.SMTP_SSL(cfg['smtp_host'], port) as s:
                s.login(cfg['smtp_user'], cfg['smtp_pass'])
                s.sendmail(cfg['smtp_user'], to_email, msg.as_string())
        else:
            with smtplib.SMTP(cfg['smtp_host'], port, timeout=15) as s:
                s.ehlo()
                s.starttls()
                s.ehlo()
                s.login(cfg['smtp_user'], cfg['smtp_pass'])
                s.sendmail(cfg['smtp_user'], to_email, msg.as_string())
        return True, 'Sent'
    except smtplib.SMTPAuthenticationError as e:
        if 'office365' in cfg['smtp_host'].lower() or 'outlook' in cfg['smtp_host'].lower():
            return False, 'Office 365 Auth fail: SMTP Auth feature enable karo → M365 Admin Center → Settings → Mail → Authenticated SMTP'
        return False, f'Authentication failed: {str(e)}'
    except Exception as e:
        return False, str(e)

# ===== SETTINGS =====
@app.route('/settings')
@login_required
def settings():
    if session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    conn = get_db()
    users    = rows(conn.execute('SELECT * FROM users ORDER BY name').fetchall())
    emp_list = rows(conn.execute('SELECT emp_id, name FROM employees WHERE status="active" ORDER BY name').fetchall())
    email_cfg= dict(conn.execute('SELECT * FROM email_settings WHERE id=1').fetchone() or {})
    loc_settings = rows(conn.execute('SELECT * FROM location_settings ORDER BY id').fetchall())
    conn.close()
    loc_settings_link = True  # flag to show location link in settings
    return render_template('settings.html', users=users, emp_list=emp_list, email_cfg=email_cfg, loc_settings=loc_settings, loc_settings_link=loc_settings_link)

@app.route('/api/users/add', methods=['POST'])
@admin_required
def api_add_user():
    data = request.json
    conn = get_db()
    try:
        conn.execute(
            'INSERT INTO users (username, password, role, name, phone, email, emp_id, can_whatsapp) VALUES (?,?,?,?,?,?,?,?)',
            (data['username'], data['password'], data.get('role','viewer'),
             data.get('name',''), data.get('phone',''), data.get('email',''),
             data.get('emp_id',''), 1 if data.get('can_whatsapp') else 0)
        )
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/users/edit', methods=['POST'])
@admin_required
def api_edit_user():
    data = request.json
    conn = get_db()
    if data.get('password'):
        conn.execute('UPDATE users SET role=?,name=?,phone=?,email=?,emp_id=?,can_whatsapp=?,password=? WHERE id=?',
            (data['role'], data.get('name',''), data.get('phone',''), data.get('email',''),
             data.get('emp_id',''), 1 if data.get('can_whatsapp') else 0, data['password'], data['id']))
    else:
        conn.execute('UPDATE users SET role=?,name=?,phone=?,email=?,emp_id=?,can_whatsapp=? WHERE id=?',
            (data['role'], data.get('name',''), data.get('phone',''), data.get('email',''),
             data.get('emp_id',''), 1 if data.get('can_whatsapp') else 0, data['id']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/users/delete', methods=['POST'])
@admin_required
def api_delete_user():
    data = request.json
    conn = get_db()
    conn.execute("DELETE FROM users WHERE id=? AND username != 'admin'", (data['id'],))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/users/change_password', methods=['POST'])
@login_required
def api_change_password():
    """User apna password change kare"""
    data = request.json
    conn = get_db()
    u = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    if not u:
        conn.close()
        return jsonify({'error': 'User not found'}), 404
    if u['password'] != data.get('old_password','').strip():
        conn.close()
        return jsonify({'error': 'Purana password galat hai'}), 400
    new_pw = data.get('new_password','').strip()
    if len(new_pw) < 4:
        conn.close()
        return jsonify({'error': 'Naya password kam se kam 4 characters ka hona chahiye'}), 400
    conn.execute('UPDATE users SET password=? WHERE id=?', (new_pw, session['user_id']))
    conn.commit()
    conn.close()
    # Send email notification
    if u['email']:
        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:480px;margin:0 auto;padding:24px;background:#f9f9f9;border-radius:12px">
          <h2 style="color:#1a3a5c;margin-bottom:8px">🔐 Password Changed</h2>
          <p style="color:#444">Namaste <b>{u['name'] or u['username']}</b>,</p>
          <p style="color:#444">Aapka HRMS password successfully change ho gaya hai.</p>
          <div style="background:#fff;border:1px solid #e0e0e0;border-radius:8px;padding:16px;margin:16px 0">
            <b>Username:</b> {u['username']}<br>
            <b>Time:</b> {datetime.now().strftime('%d %b %Y, %I:%M %p')}
          </div>
          <p style="color:#888;font-size:12px">Agar aapne yeh change nahi kiya toh turant admin se contact karein.</p>
          <p style="color:#888;font-size:12px">— EasySource HRMS</p>
        </div>"""
        send_email(u['email'], '🔐 HRMS Password Changed', html)
    return jsonify({'success': True})

@app.route('/api/email_settings/save', methods=['POST'])
@admin_required
def api_save_email_settings():
    data = request.json
    conn = get_db()
    conn.execute('''UPDATE email_settings SET smtp_host=?,smtp_port=?,smtp_user=?,smtp_pass=?,from_name=?,enabled=? WHERE id=1''',
        (data.get('smtp_host','smtp.gmail.com'), int(data.get('smtp_port',587)),
         data.get('smtp_user',''), data.get('smtp_pass',''),
         data.get('from_name','EasySource HRMS'), 1 if data.get('enabled') else 0))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/email_settings/test', methods=['POST'])
@admin_required
def api_test_email():
    data = request.json
    ok, msg = send_email(data.get('to_email',''), '✅ HRMS Email Test', '<h3>Email working hai! ✅</h3><p>EasySource HRMS se test email.</p>')
    return jsonify({'success': ok, 'message': msg})


# ===== ROSTER RULES (Employee-wise Default Shifts) =====

@app.route('/api/roster/rules/get', methods=['GET'])
@login_required
def api_roster_rules_get():
    conn = get_db()
    rules = rows(conn.execute('''
        SELECT rr.*, e.name, e.department, e.designation, e.office_setting
        FROM roster_rules rr
        JOIN employees e ON rr.emp_id = e.emp_id
        WHERE e.status='active'
        ORDER BY e.name
    ''').fetchall())
    conn.close()
    return jsonify(rules)

@app.route('/api/roster/rules/save', methods=['POST'])
@login_required
def api_roster_rules_save():
    data   = request.json
    emp_id = data.get('emp_id','').strip()
    if not emp_id:
        return jsonify({'success': False, 'error': 'emp_id required'})
    conn = get_db()
    existing = conn.execute('SELECT id FROM roster_rules WHERE emp_id=?', (emp_id,)).fetchone()
    days = ('mon','tue','wed','thu','fri','sat','sun')
    vals = tuple(data.get(d,'W') for d in days)
    rule_name = data.get('rule_name','').strip()
    notes     = data.get('notes','').strip()
    if existing:
        conn.execute(
            'UPDATE roster_rules SET mon=?,tue=?,wed=?,thu=?,fri=?,sat=?,sun=?,rule_name=?,notes=?,created_by=?,updated_at=CURRENT_TIMESTAMP WHERE emp_id=?',
            vals + (rule_name, notes, session['username'], emp_id)
        )
    else:
        conn.execute(
            'INSERT INTO roster_rules (emp_id,mon,tue,wed,thu,fri,sat,sun,rule_name,notes,created_by) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
            (emp_id,) + vals + (rule_name, notes, session['username'])
        )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/roster/rules/delete', methods=['POST'])
@admin_required
def api_roster_rules_delete():
    data = request.json
    conn = get_db()
    conn.execute('DELETE FROM roster_rules WHERE emp_id=?', (data['emp_id'],))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/roster/rules/bulk-save', methods=['POST'])
@login_required
def api_roster_rules_bulk_save():
    data    = request.json
    emp_ids = data.get('emp_ids', [])
    if not emp_ids:
        return jsonify({'success': False, 'error': 'No employees selected'})
    days      = ('mon','tue','wed','thu','fri','sat','sun')
    rule_data = tuple(data.get(d,'W') for d in days)
    rule_name = data.get('rule_name','').strip()
    notes     = data.get('notes','').strip()
    conn  = get_db()
    saved = 0
    for eid in emp_ids:
        existing = conn.execute('SELECT id FROM roster_rules WHERE emp_id=?', (eid,)).fetchone()
        if existing:
            conn.execute(
                'UPDATE roster_rules SET mon=?,tue=?,wed=?,thu=?,fri=?,sat=?,sun=?,rule_name=?,notes=?,created_by=?,updated_at=CURRENT_TIMESTAMP WHERE emp_id=?',
                rule_data + (rule_name, notes, session['username'], eid)
            )
        else:
            conn.execute(
                'INSERT INTO roster_rules (emp_id,mon,tue,wed,thu,fri,sat,sun,rule_name,notes,created_by) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                (eid,) + rule_data + (rule_name, notes, session['username'])
            )
        saved += 1
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'saved': saved})

@app.route('/api/roster/rules/apply', methods=['POST'])
@login_required
def api_roster_rules_apply():
    """Apply employee default rules to a specific week's roster"""
    data       = request.json
    week_start = data.get('week_start')
    emp_ids    = data.get('emp_ids', [])
    overwrite  = data.get('overwrite', True)
    if not week_start:
        return jsonify({'success': False, 'error': 'week_start required'})
    conn = get_db()
    if emp_ids:
        ph    = ','.join(['?']*len(emp_ids))
        rules = rows(conn.execute(f'SELECT * FROM roster_rules WHERE emp_id IN ({ph})', emp_ids).fetchall())
    else:
        rules = rows(conn.execute('SELECT * FROM roster_rules').fetchall())
    saved = skipped = 0
    for rule in rules:
        eid      = rule['emp_id']
        existing = conn.execute('SELECT id FROM roster WHERE emp_id=? AND week_start=?', (eid, week_start)).fetchone()
        if existing and not overwrite:
            skipped += 1
            continue
        if existing:
            conn.execute(
                'UPDATE roster SET mon=?,tue=?,wed=?,thu=?,fri=?,sat=?,sun=? WHERE emp_id=? AND week_start=?',
                (rule['mon'],rule['tue'],rule['wed'],rule['thu'],rule['fri'],rule['sat'],rule['sun'], eid, week_start)
            )
        else:
            conn.execute(
                'INSERT INTO roster (emp_id,week_start,mon,tue,wed,thu,fri,sat,sun) VALUES (?,?,?,?,?,?,?,?,?)',
                (eid, week_start, rule['mon'],rule['tue'],rule['wed'],rule['thu'],rule['fri'],rule['sat'],rule['sun'])
            )
        saved += 1
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'saved': saved, 'skipped': skipped})

@app.route('/api/employees/list', methods=['GET','POST'])
@login_required
def api_employees_list():
    conn  = get_db()
    emps  = rows(conn.execute('SELECT emp_id,name,department,designation,office_setting,status FROM employees WHERE status="active" ORDER BY name').fetchall())
    conn.close()
    return jsonify(emps)


# ===== ROSTER TEMPLATES =====

@app.route('/api/roster/templates/get', methods=['GET'])
@login_required
def api_roster_templates_get():
    try:
        conn = get_db()
        seed_default_templates(conn)
        templates = rows(conn.execute('SELECT * FROM roster_templates ORDER BY name').fetchall())
        for t in templates:
            try:
                t['mapped_count'] = conn.execute(
                    'SELECT COUNT(*) FROM roster_rules WHERE rule_name=?', (t['name'],)
                ).fetchone()[0]
            except:
                t['mapped_count'] = 0
        conn.close()
        return jsonify(templates)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/roster/templates/save', methods=['POST'])
@login_required
def api_roster_templates_save():
    data = request.json
    name = data.get('name','').strip()
    if not name:
        return jsonify({'success': False, 'error': 'Template name required'})
    tid   = data.get('id')
    days  = ('mon','tue','wed','thu','fri','sat','sun')
    vals  = tuple(data.get(d,'W') for d in days)
    desc  = data.get('description','').strip()
    color = data.get('color','primary')
    conn  = get_db()
    if tid:
        conn.execute(
            'UPDATE roster_templates SET name=?,description=?,color=?,mon=?,tue=?,wed=?,thu=?,fri=?,sat=?,sun=? WHERE id=?',
            (name, desc, color) + vals + (tid,)
        )
    else:
        conn.execute(
            'INSERT INTO roster_templates (name,description,color,mon,tue,wed,thu,fri,sat,sun,created_by) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
            (name, desc, color) + vals + (session['username'],)
        )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/roster/templates/delete', methods=['POST'])
@login_required
def api_roster_templates_delete():
    data = request.json
    tid  = data.get('id')
    conn = get_db()
    # Get template name first to clean up rules
    t = conn.execute('SELECT name FROM roster_templates WHERE id=?', (tid,)).fetchone()
    if t:
        conn.execute('UPDATE roster_rules SET rule_name="" WHERE rule_name=?', (t['name'],))
    conn.execute('DELETE FROM roster_templates WHERE id=?', (tid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})



@app.route('/api/roster/report/download', methods=['GET','POST'])
@login_required
def api_roster_report_download():
    """Generate and download Excel roster report"""
    import calendar, sqlite3
    from datetime import date, timedelta
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io

    if request.method == 'GET':
        year  = int(request.args.get('year', date.today().year))
        month = int(request.args.get('month', date.today().month))
    else:
        data  = request.json or {}
        year  = int(data.get('year', date.today().year))
        month = int(data.get('month', date.today().month))

    SHIFT_COLORS = {
        'W':'FF2ECC71','WO':'FF95A5A6','H':'FF1ABC9C',
        'L':'FFF39C12','OD':'FFFD7E14','WFH':'FF3D5AFE','HD':'FF9B59B6',
    }
    def hfill(c): return PatternFill('solid',start_color=c,end_color=c)
    def tborder():
        s=Side(style='thin',color='FFD0D0D0')
        return Border(left=s,right=s,top=s,bottom=s)

    conn2 = get_db()
    emps = rows(conn2.execute("SELECT emp_id,name,department,office_setting FROM employees WHERE status='active' ORDER BY department,name").fetchall())
    _,days_in_month = calendar.monthrange(year,month)
    all_dates = [date(year,month,d) for d in range(1,days_in_month+1)]

    roster_rows2 = conn2.execute("SELECT emp_id,week_start,mon,tue,wed,thu,fri,sat,sun FROM roster WHERE week_start BETWEEN ? AND ?",
        (str(date(year,month,1) - timedelta(days=7)), str(date(year,month,days_in_month) + timedelta(days=7)))
    ).fetchall()
    conn2.close()

    DAY_KEYS2 = ['mon','tue','wed','thu','fri','sat','sun']
    emp_roster2 = {}
    for r2 in roster_rows2:
        eid=r2['emp_id']; ws2=date.fromisoformat(r2['week_start'])
        if eid not in emp_roster2: emp_roster2[eid]={}
        for i,dk in enumerate(DAY_KEYS2):
            d=ws2+timedelta(days=i)
            if d.month==month: emp_roster2[eid][d]=r2[dk] or 'W'

    wb2=Workbook()
    DAY_ABBR2=['Mo','Tu','We','Th','Fr','Sa','Su']
    MONTH_NAME=calendar.month_name[month]

    # Sheet 1 - Monthly Roster
    ws_r=wb2.active; ws_r.title='Monthly Roster'
    ws_r.sheet_view.showGridLines=False; ws_r.freeze_panes='F3'

    ws_r.merge_cells(f"A1:{get_column_letter(5+days_in_month+5)}1")
    t=ws_r['A1']; t.value=f'EasySource HRMS — Roster Report  |  {MONTH_NAME} {year}'
    t.font=Font(name='Arial',size=13,bold=True,color='FFFFFFFF')
    t.fill=hfill('FF1A1F2E'); t.alignment=Alignment(horizontal='center',vertical='center')
    ws_r.row_dimensions[1].height=28

    for ci,h in enumerate(['#','Emp ID','Name','Department','Office'],1):
        c=ws_r.cell(row=2,column=ci,value=h)
        c.font=Font(name='Arial',size=9,bold=True,color='FFFFFFFF')
        c.fill=hfill('FF0D3349'); c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tborder()

    for di,d in enumerate(all_dates):
        col=6+di; is_sat=d.weekday()==5; is_sun=d.weekday()==6
        bg='FF1E4D2B' if is_sat else ('FF4D1E1E' if is_sun else 'FF0F4D6A')
        c=ws_r.cell(row=2,column=col,value=f'{d.day}\n{DAY_ABBR2[d.weekday()]}')
        c.font=Font(name='Arial',size=8,bold=True,color='FFFFFFFF')
        c.fill=hfill(bg); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=tborder()
    ws_r.row_dimensions[2].height=30

    for si,sc in enumerate(['W Days','WO','WFH','OD','Leave']):
        col=6+days_in_month+si; c=ws_r.cell(row=2,column=col,value=sc)
        c.font=Font(name='Arial',size=8,bold=True,color='FFFFFFFF')
        c.fill=hfill('FF1A3A5C'); c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tborder()

    for ri2,emp in enumerate(emps):
        row=ri2+3; eid=emp['emp_id']; shifts=emp_roster2.get(eid,{})
        bg_row='FF1C2333' if ri2%2==0 else 'FF161B27'
        def bc(col,val,bold=False,align='left'):
            c=ws_r.cell(row=row,column=col,value=val)
            c.font=Font(name='Arial',size=9,bold=bold,color='FFE6EDF3')
            c.fill=hfill(bg_row); c.alignment=Alignment(horizontal=align,vertical='center'); c.border=tborder()
        bc(1,ri2+1,align='center'); bc(2,eid,align='center'); bc(3,emp['name'],bold=True)
        bc(4,emp.get('department','') or ''); bc(5,emp.get('office_setting','') or '')
        w_c=wo_c=wfh_c=od_c=lv_c=0
        for di2,d in enumerate(all_dates):
            shift=shifts.get(d,'')
            sc2=ws_r.cell(row=row,column=6+di2,value=shift or '')
            col2=SHIFT_COLORS.get(shift,'FFEEEEEE')
            sc2.fill=PatternFill('solid',start_color=col2,end_color=col2)
            sc2.font=Font(name='Arial',size=8,bold=True,color='FFFFFFFF')
            sc2.alignment=Alignment(horizontal='center',vertical='center'); sc2.border=tborder()
            if shift=='W': w_c+=1
            elif shift=='WO': wo_c+=1
            elif shift=='WFH': wfh_c+=1
            elif shift=='OD': od_c+=1
            elif shift in('L','H','HD'): lv_c+=1
        for si2,val in enumerate([w_c,wo_c,wfh_c,od_c,lv_c]):
            col3=6+days_in_month+si2; c=ws_r.cell(row=row,column=col3,value=val)
            c.font=Font(name='Arial',size=9,bold=True,color='FF00C896' if si2==0 else 'FFE6EDF3')
            c.fill=hfill(bg_row); c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tborder()
        ws_r.row_dimensions[row].height=18

    ws_r.column_dimensions['A'].width=4; ws_r.column_dimensions['B'].width=9
    ws_r.column_dimensions['C'].width=22; ws_r.column_dimensions['D'].width=14; ws_r.column_dimensions['E'].width=16
    for di3 in range(days_in_month): ws_r.column_dimensions[get_column_letter(6+di3)].width=4.2
    for si3 in range(5): ws_r.column_dimensions[get_column_letter(6+days_in_month+si3)].width=7

    # Sheet 2 - Summary
    ws_s=wb2.create_sheet('Employee Summary')
    ws_s.sheet_view.showGridLines=False; ws_s.freeze_panes='A3'
    ws_s.merge_cells('A1:J1'); t2=ws_s['A1']
    t2.value=f'Employee Shift Summary  |  {MONTH_NAME} {year}'
    t2.font=Font(name='Arial',size=13,bold=True,color='FFFFFFFF')
    t2.fill=hfill('FF1A1F2E'); t2.alignment=Alignment(horizontal='center',vertical='center')
    ws_s.row_dimensions[1].height=28
    for ci,h in enumerate(['#','Emp ID','Name','Department','Office','Working (W)','Week Off','WFH','OD','Leave'],1):
        c=ws_s.cell(row=2,column=ci,value=h)
        c.font=Font(name='Arial',size=9,bold=True,color='FFFFFFFF')
        c.fill=hfill('FF0D3349'); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=tborder()
    ws_s.row_dimensions[2].height=30
    for ri3,emp in enumerate(emps):
        row=ri3+3; eid=emp['emp_id']; shifts=emp_roster2.get(eid,{})
        bg_row='FF1C2333' if ri3%2==0 else 'FF161B27'
        cnt={'W':0,'WO':0,'WFH':0,'OD':0,'lv':0}
        for d in all_dates:
            s=shifts.get(d,'')
            if s=='W': cnt['W']+=1
            elif s=='WO': cnt['WO']+=1
            elif s=='WFH': cnt['WFH']+=1
            elif s=='OD': cnt['OD']+=1
            elif s in('L','H','HD'): cnt['lv']+=1
        for ci2,v in enumerate([ri3+1,eid,emp['name'],emp.get('department','') or '',emp.get('office_setting','') or '',cnt['W'],cnt['WO'],cnt['WFH'],cnt['OD'],cnt['lv']],1):
            c=ws_s.cell(row=row,column=ci2,value=v)
            c.font=Font(name='Arial',size=9,bold=(ci2==3),color='FF00C896' if ci2==6 else 'FFE6EDF3')
            c.fill=hfill(bg_row); c.alignment=Alignment(horizontal='center' if ci2>=6 or ci2<=2 else 'left',vertical='center'); c.border=tborder()
    for i4,w4 in enumerate([4,9,24,16,18,13,13,10,10,13]):
        ws_s.column_dimensions[get_column_letter(i4+1)].width=w4

    buf=_io.BytesIO(); wb2.save(buf); buf.seek(0)
    fname=f'Roster_Report_{MONTH_NAME}_{year}.xlsx'
    from flask import make_response
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
    resp.headers['Content-Transfer-Encoding'] = 'binary'
    resp.headers['Cache-Control'] = 'no-cache'
    resp.headers['X-Content-Type-Options'] = 'nosniff'
    return resp



# ===== ROSTER WHATSAPP NOTIFY =====
@app.route('/api/roster/notify/preview', methods=['POST'])
@login_required
def api_roster_notify_preview():
    """Preview roster message for an employee for a given week"""
    from datetime import date, timedelta
    data = request.json
    emp_id = data.get('emp_id')
    week_start = data.get('week_start')
    conn = get_db()
    emp = row(conn.execute('SELECT name, phone FROM employees WHERE emp_id=?', (emp_id,)).fetchone())
    if not emp:
        conn.close()
        return jsonify({'success': False, 'error': 'Employee not found'})
    roster_row = row(conn.execute('SELECT * FROM roster WHERE emp_id=? AND week_start=?', (emp_id, week_start)).fetchone())
    conn.close()
    if not roster_row:
        return jsonify({'success': False, 'error': 'No roster found for this week'})
    day_labels = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
    day_keys   = ['mon','tue','wed','thu','fri','sat','sun']
    ws = date.fromisoformat(week_start)
    lines = [f"📋 *Roster Schedule*", f"👤 {emp['name']}", f"📅 Week: {ws.strftime('%d %b')} - {(ws+timedelta(days=6)).strftime('%d %b %Y')}", ""]
    shift_emoji = {'W':'✅','WO':'🔴','WFH':'🏠','OD':'🏢','L':'🌴','H':'🎉','HD':'⚡'}
    for i, dk in enumerate(day_keys):
        d = ws + timedelta(days=i)
        shift = roster_row.get(dk, '')
        emoji = shift_emoji.get(shift, '▪️')
        lines.append(f"{emoji} {day_labels[i]} ({d.strftime('%d %b')}): *{shift}*")
    lines += ["", "— EasySource HRMS"]
    message = "\n".join(lines)
    import urllib.parse
    phone = (emp.get('phone') or '').strip().replace('+','').replace(' ','').replace('-','')
    wa_link = f"https://wa.me/91{phone}?text={urllib.parse.quote(message)}" if phone else ''
    return jsonify({'success': True, 'message': message, 'wa_link': wa_link, 'phone': phone, 'name': emp['name']})

@app.route('/api/roster/notify/bulk-preview', methods=['POST'])
@login_required
def api_roster_notify_bulk():
    """Get wa.me links for all/filtered employees for a week"""
    from datetime import date, timedelta
    import urllib.parse
    data = request.json
    week_start = data.get('week_start')
    office_filter = data.get('office', '')
    conn = get_db()
    query = "SELECT e.emp_id, e.name, e.phone, e.office_setting FROM employees e WHERE e.status='active' AND e.phone IS NOT NULL AND e.phone != ''"
    params = []
    if office_filter:
        query += " AND e.office_setting=?"
        params.append(office_filter)
    emps = rows(conn.execute(query, params).fetchall())
    day_labels = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
    day_keys   = ['mon','tue','wed','thu','fri','sat','sun']
    shift_emoji = {'W':'✅','WO':'🔴','WFH':'🏠','OD':'🏢','L':'🌴','H':'🎉','HD':'⚡'}
    ws = date.fromisoformat(week_start)
    result = []
    for emp in emps:
        eid = emp['emp_id']
        r2 = row(conn.execute('SELECT * FROM roster WHERE emp_id=? AND week_start=?', (eid, week_start)).fetchone())
        if not r2:
            continue
        lines = [f"📋 *Roster Schedule*", f"👤 {emp['name']}", f"📅 Week: {ws.strftime('%d %b')} - {(ws+timedelta(days=6)).strftime('%d %b %Y')}", ""]
        for i, dk in enumerate(day_keys):
            d = ws + timedelta(days=i)
            shift = r2.get(dk, '')
            emoji = shift_emoji.get(shift, '▪️')
            lines.append(f"{emoji} {day_labels[i]} ({d.strftime('%d %b')}): *{shift}*")
        lines += ["", "— EasySource HRMS"]
        message = "\n".join(lines)
        phone = (emp.get('phone') or '').strip().replace('+','').replace(' ','').replace('-','')
        wa_link = f"https://wa.me/91{phone}?text={urllib.parse.quote(message)}" if phone else ''
        result.append({'emp_id': eid, 'name': emp['name'], 'phone': phone, 'wa_link': wa_link, 'office': emp.get('office_setting','')})
    conn.close()
    return jsonify({'success': True, 'employees': result, 'week_start': week_start})


# ===== LOCATION SETTINGS =====
@app.route('/api/location/settings/get', methods=['GET'])
@login_required
def api_location_settings_get():
    conn = get_db()
    locs = rows(conn.execute('SELECT * FROM location_settings ORDER BY id').fetchall())
    conn.close()
    return jsonify(locs)

@app.route('/api/location/settings/save', methods=['POST'])
@admin_required
def api_location_settings_save():
    data = request.json
    lid  = data.get('id')
    conn = get_db()
    if lid:
        conn.execute('UPDATE location_settings SET office_name=?,latitude=?,longitude=?,radius_meters=?,is_active=? WHERE id=?',
            (data['office_name'], data['latitude'], data['longitude'], data.get('radius_meters',200), data.get('is_active',1), lid))
    else:
        conn.execute('INSERT INTO location_settings (office_name,latitude,longitude,radius_meters,is_active,created_by) VALUES (?,?,?,?,?,?)',
            (data['office_name'], data['latitude'], data['longitude'], data.get('radius_meters',200), data.get('is_active',1), session['username']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/location/settings/delete', methods=['POST'])
@admin_required
def api_location_settings_delete():
    data = request.json
    conn = get_db()
    conn.execute('DELETE FROM location_settings WHERE id=?', (data['id'],))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/location-settings')
@login_required
def location_settings_page():
    if session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    conn = get_db()
    locs = rows(conn.execute('SELECT * FROM location_settings ORDER BY id').fetchall())
    conn.close()
    return render_template_string(LOCATION_SETTINGS_TEMPLATE, locs=locs)

@app.route('/employee-location')
@login_required
def employee_location_page():
    if session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    return render_template_string(EMPLOYEE_LOCATION_TEMPLATE)

LOCATION_SETTINGS_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Location Settings — EasySource HRMS</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css" rel="stylesheet">
<style>
  :root{--bg:#0d1117;--card:#161b27;--card2:#1c2333;--border:rgba(255,255,255,.08);--accent:#00c896;--text:#e6edf3;--muted:#8b949e}
  body{background:var(--bg);color:var(--text);font-family:Arial,sans-serif;min-height:100vh;padding:24px}
  .page-card{background:var(--card);border:1px solid var(--border);border-radius:14px;padding:24px;max-width:900px;margin:0 auto}
  .page-hdr{display:flex;align-items:center;justify-content:space-between;margin-bottom:24px}
  .page-title{font-size:20px;font-weight:700;color:var(--accent)}
  .back-btn{background:rgba(255,255,255,.07);border:1px solid var(--border);color:var(--text);border-radius:8px;padding:6px 14px;text-decoration:none;font-size:13px}
  .back-btn:hover{background:rgba(255,255,255,.12);color:var(--text)}
  table{width:100%;border-collapse:collapse}
  th{background:#0d1f35;color:var(--muted);font-size:11px;letter-spacing:.5px;text-transform:uppercase;padding:10px 12px;border-bottom:1px solid var(--border)}
  td{padding:10px 12px;border-bottom:1px solid var(--border);vertical-align:middle}
  tr:hover td{background:rgba(255,255,255,.02)}
  input[type=text],input[type=number]{background:#0d1117;border:1px solid var(--border);color:var(--text);border-radius:6px;padding:5px 10px;font-size:13px;width:100%}
  input[type=text]:focus,input[type=number]:focus{outline:none;border-color:var(--accent)}
  .btn-save{background:linear-gradient(135deg,#00c896,#00a87a);color:#fff;border:none;border-radius:6px;padding:5px 14px;font-size:12px;cursor:pointer;font-weight:600}
  .btn-del{background:rgba(239,68,68,.15);border:1px solid rgba(239,68,68,.3);color:#ef4444;border-radius:6px;padding:5px 10px;font-size:12px;cursor:pointer}
  .btn-map{background:rgba(61,90,254,.15);border:1px solid rgba(61,90,254,.3);color:#3d5afe;border-radius:6px;padding:5px 10px;font-size:12px;text-decoration:none}
  .btn-add{background:linear-gradient(135deg,#00c896,#00a87a);color:#fff;border:none;border-radius:8px;padding:8px 18px;font-size:13px;font-weight:700;cursor:pointer}
  .badge-active{background:rgba(0,200,150,.15);color:#00c896;border:1px solid rgba(0,200,150,.3);border-radius:12px;padding:2px 8px;font-size:11px}
  .badge-inactive{background:rgba(150,150,150,.15);color:#8b949e;border:1px solid rgba(150,150,150,.3);border-radius:12px;padding:2px 8px;font-size:11px}
  .info-box{background:rgba(0,200,150,.07);border:1px solid rgba(0,200,150,.2);border-radius:10px;padding:14px;font-size:13px;color:var(--muted);margin-bottom:20px}
  .toast-msg{position:fixed;top:20px;right:20px;background:#00c896;color:#fff;padding:10px 20px;border-radius:8px;font-weight:600;display:none;z-index:9999}
</style>
</head>
<body>
<div class="page-card">
  <div class="page-hdr">
    <div class="page-title">📍 Attendance Location Range Settings</div>
    <div class="d-flex gap-2">
      <button class="btn-add" onclick="addRow()">+ Add Office</button>
      <a href="/settings" class="back-btn">← Settings</a>
    </div>
  </div>

  <div class="info-box">
    💡 <b style="color:var(--text)">Kaise use karein:</b><br>
    Har office/location ke liye <b>latitude, longitude aur radius</b> set karo.<br>
    Employee jab attendance mark kare aur uski location is range ke andar ho → <b style="color:#00c896">✅ In Range</b><br>
    Bahar ho → <b style="color:#ef4444">❌ Out of Range</b> (attendance tab bhi mark hogi, sirf flag lagega)<br><br>
    📌 Google Maps pe jaao → apni office pe right-click → <b>"What's here?"</b> → coordinates copy karo
  </div>

  <table>
    <thead>
      <tr>
        <th style="width:22%">Office Name</th>
        <th style="width:16%">Latitude</th>
        <th style="width:16%">Longitude</th>
        <th style="width:14%">Radius (meters)</th>
        <th style="width:8%">Active</th>
        <th style="width:24%">Actions</th>
      </tr>
    </thead>
    <tbody id="locBody">
      {% for loc in locs %}
      <tr id="row_{{ loc.id }}">
        <td><input type="text" id="n_{{ loc.id }}" value="{{ loc.office_name }}"></td>
        <td><input type="number" step="0.000001" id="lat_{{ loc.id }}" value="{{ loc.latitude }}"></td>
        <td><input type="number" step="0.000001" id="lng_{{ loc.id }}" value="{{ loc.longitude }}"></td>
        <td><input type="number" id="rad_{{ loc.id }}" value="{{ loc.radius_meters }}"></td>
        <td style="text-align:center">
          <input type="checkbox" id="act_{{ loc.id }}" {% if loc.is_active %}checked{% endif %} style="width:16px;height:16px;cursor:pointer">
        </td>
        <td>
          <button class="btn-save me-1" onclick="save({{ loc.id }})">💾 Save</button>
          <button class="btn-del me-1" onclick="del({{ loc.id }})">🗑</button>
          <a class="btn-map" href="https://maps.google.com?q={{ loc.latitude }},{{ loc.longitude }}" target="_blank">🗺 Map</a>
        </td>
      </tr>
      {% endfor %}
    </tbody>
  </table>

  {% if not locs %}
  <div style="text-align:center;padding:30px;color:var(--muted)">Koi location set nahi hai. "+ Add Office" karo.</div>
  {% endif %}
</div>

<div class="toast-msg" id="toast"></div>

<script>
function showToast(msg, ok=true){
  var t=document.getElementById('toast');
  t.textContent=msg;
  t.style.background=ok?'#00c896':'#ef4444';
  t.style.display='block';
  setTimeout(()=>t.style.display='none',2500);
}

async function save(lid){
  var data={
    id:lid,
    office_name:document.getElementById('n_'+lid).value.trim(),
    latitude:parseFloat(document.getElementById('lat_'+lid).value),
    longitude:parseFloat(document.getElementById('lng_'+lid).value),
    radius_meters:parseInt(document.getElementById('rad_'+lid).value)||200,
    is_active:document.getElementById('act_'+lid).checked?1:0
  };
  if(!data.office_name||!data.latitude||!data.longitude){showToast('Sab fields fill karo!',false);return;}
  var res=await fetch('/api/location/settings/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(data)});
  var d=await res.json();
  d.success?showToast('✅ Saved: '+data.office_name):showToast('❌ Error: '+(d.error||''),false);
}

async function del(lid){
  if(!confirm('Delete this location?'))return;
  var res=await fetch('/api/location/settings/delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id:lid})});
  var d=await res.json();
  if(d.success){document.getElementById('row_'+lid).remove();showToast('Deleted!');}
}

var nc=9000;
function addRow(){
  var lid=nc++;
  var tr=document.createElement('tr');
  tr.id='row_'+lid;
  tr.innerHTML=`
    <td><input type="text" id="n_${lid}" placeholder="e.g. Head Office Delhi"></td>
    <td><input type="number" step="0.000001" id="lat_${lid}" placeholder="28.6139"></td>
    <td><input type="number" step="0.000001" id="lng_${lid}" placeholder="77.2090"></td>
    <td><input type="number" id="rad_${lid}" value="200"></td>
    <td style="text-align:center"><input type="checkbox" id="act_${lid}" checked style="width:16px;height:16px;cursor:pointer"></td>
    <td>
      <button class="btn-save me-1" onclick="saveNew(${lid})">💾 Save</button>
      <button class="btn-del" onclick="this.closest('tr').remove()">✕</button>
    </td>`;
  document.getElementById('locBody').appendChild(tr);
  tr.querySelector('input[type=text]').focus();
}

async function saveNew(lid){
  var data={
    office_name:document.getElementById('n_'+lid).value.trim(),
    latitude:parseFloat(document.getElementById('lat_'+lid).value),
    longitude:parseFloat(document.getElementById('lng_'+lid).value),
    radius_meters:parseInt(document.getElementById('rad_'+lid).value)||200,
    is_active:document.getElementById('act_'+lid).checked?1:0
  };
  if(!data.office_name||!data.latitude||!data.longitude){showToast('Sab fields fill karo!',false);return;}
  var res=await fetch('/api/location/settings/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(data)});
  var d=await res.json();
  if(d.success){showToast('✅ Saved! Reloading...');setTimeout(()=>location.reload(),1000);}
  else showToast('❌ Error: '+(d.error||''),false);
}
</script>
</body>
</html>
"""

EMPLOYEE_LOCATION_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Employee Location Settings — EasySource HRMS</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<style>
:root{--bg:#0d1117;--card:#161b27;--card2:#1c2333;--border:rgba(255,255,255,.08);--accent:#00c896;--text:#e6edf3;--muted:#8b949e}
body{background:var(--bg);color:var(--text);font-family:Arial,sans-serif;padding:20px}
.page-card{background:var(--card);border:1px solid var(--border);border-radius:14px;padding:24px;max-width:1150px;margin:0 auto}
.page-title{font-size:20px;font-weight:700;color:var(--accent);margin-bottom:4px}
.back-btn{background:rgba(255,255,255,.07);border:1px solid var(--border);color:var(--text);border-radius:8px;padding:6px 14px;text-decoration:none;font-size:13px}
.back-btn:hover{background:rgba(255,255,255,.12);color:var(--text)}
th{background:#0d1f35;color:var(--muted);font-size:11px;letter-spacing:.5px;text-transform:uppercase;padding:10px 12px;border-bottom:1px solid var(--border);white-space:nowrap}
td{padding:8px 10px;border-bottom:1px solid var(--border);vertical-align:middle;font-size:13px}
tr:hover td{background:rgba(255,255,255,.02)}
.tbl-input{background:#0d1117;border:1px solid var(--border);color:var(--text);border-radius:6px;padding:4px 8px;font-size:12px;width:100%}
.tbl-select{background:#0d1117;border:1px solid var(--border);color:var(--text);border-radius:6px;padding:4px 6px;font-size:12px;width:130px}
.btn-save{background:linear-gradient(135deg,#00c896,#00a87a);color:#fff;border:none;border-radius:6px;padding:5px 14px;font-size:12px;cursor:pointer;font-weight:600}
.badge-office{background:rgba(0,200,150,.15);color:#00c896;border:1px solid rgba(0,200,150,.3);border-radius:10px;padding:2px 8px;font-size:11px;font-weight:600}
.badge-home{background:rgba(61,90,254,.15);color:#7c9fff;border:1px solid rgba(61,90,254,.3);border-radius:10px;padding:2px 8px;font-size:11px;font-weight:600}
.badge-any{background:rgba(255,152,0,.15);color:#ffb74d;border:1px solid rgba(255,152,0,.3);border-radius:10px;padding:2px 8px;font-size:11px;font-weight:600}
.info-box{background:rgba(0,200,150,.07);border:1px solid rgba(0,200,150,.2);border-radius:10px;padding:14px;font-size:13px;color:var(--muted);margin-bottom:20px;line-height:2}
.bulk-bar{background:var(--card2);border:1px solid var(--border);border-radius:10px;padding:12px 16px;margin-bottom:16px;display:flex;gap:12px;align-items:center;flex-wrap:wrap}
.filter-row{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:16px;align-items:center}
.f-input{background:#0d1117;border:1px solid var(--border);color:var(--text);border-radius:8px;padding:6px 12px;font-size:13px}
.toast-msg{position:fixed;top:20px;right:20px;padding:10px 20px;border-radius:8px;font-weight:600;display:none;z-index:9999;color:#fff}
</style>
</head>
<body>
<div class="page-card">

  <!-- Header -->
  <div class="d-flex justify-content-between align-items-start mb-3">
    <div>
      <div class="page-title">👤 Employee-wise Location Settings</div>
      <div style="font-size:13px;color:var(--muted)">Har employee ka location mode aur WFH location set karo</div>
    </div>
    <div class="d-flex gap-2">
      <a href="/location-settings" class="back-btn">🏢 Office Locations</a>
      <a href="/settings" class="back-btn">← Settings</a>
    </div>
  </div>

  <!-- Info box -->
  <div class="info-box">
    🟢 <b style="color:#00c896">Office</b> — Sirf office range ke andar mark kar sakta hai<br>
    🔵 <b style="color:#7c9fff">Home / WFH</b> — Sirf ghar ke coordinates ke andar mark kar sakta hai (neeche set karo)<br>
    🟡 <b style="color:#ffb74d">Field / Any</b> — Koi restriction nahi, kahin se bhi mark kar sakta hai<br>
    <span style="font-size:12px">📌 <b>Home coordinates:</b> Employee Google Maps pe apna ghar dhundhe → right-click → coordinates copy kare aur yahan paste karo</span>
  </div>

  <!-- Bulk action -->
  <div class="bulk-bar">
    <b style="font-size:13px">Bulk:</b>
    <select id="bulkMode" class="f-input" style="width:160px">
      <option value="office">🏢 Office</option>
      <option value="home">🏠 Home/WFH</option>
      <option value="any">🌍 Field/Any</option>
    </select>
    <button class="btn-save" onclick="bulkApply()">✅ Apply to Selected</button>
    <span style="font-size:12px;color:var(--muted)" id="selCount">0 selected</span>
  </div>

  <!-- Filters -->
  <div class="filter-row">
    <input class="f-input" id="srch" placeholder="🔍 Name / Emp ID..." oninput="filterTable()" style="width:200px">
    <select class="f-input" id="modeFilter" onchange="filterTable()" style="width:150px">
      <option value="">All Modes</option>
      <option value="office">🏢 Office</option>
      <option value="home">🏠 Home/WFH</option>
      <option value="any">🌍 Field/Any</option>
    </select>
    <select class="f-input" id="officeFilter" onchange="filterTable()" style="width:190px">
      <option value="">All Offices</option>
      <option>Head Office Delhi</option>
      <option>Branch Office Gurgaon</option>
      <option>Work From Home</option>
      <option>Client Side</option>
      <option>Field</option>
    </select>
    <span style="font-size:12px;color:var(--muted)" id="showCount"></span>
  </div>

  <!-- Table -->
  <div style="overflow-x:auto">
    <table style="width:100%;border-collapse:collapse" id="empTable">
      <thead>
        <tr>
          <th style="width:36px"><input type="checkbox" id="selAll" onchange="toggleAll(this)" style="width:16px;height:16px;cursor:pointer"></th>
          <th>Emp ID</th>
          <th>Name</th>
          <th>Office Setting</th>
          <th>Mode</th>
          <th>Home Lat</th>
          <th>Home Lng</th>
          <th>Radius (m)</th>
          <th>Save</th>
        </tr>
      </thead>
      <tbody id="empBody">
        <tr><td colspan="9" style="text-align:center;padding:30px;color:var(--muted)">⏳ Loading...</td></tr>
      </tbody>
    </table>
  </div>
</div>

<div class="toast-msg" id="toast"></div>

<script>
var allEmps = [];

function showToast(msg, ok){
  var t = document.getElementById('toast');
  t.textContent = msg;
  t.style.background = (ok===false) ? '#ef4444' : '#00c896';
  t.style.display = 'block';
  setTimeout(function(){ t.style.display='none'; }, 2500);
}

async function loadEmps(){
  try {
    var res = await fetch('/api/employees/location/list');
    allEmps  = await res.json();
    filterTable();
  } catch(e) {
    document.getElementById('empBody').innerHTML = '<tr><td colspan="9" style="text-align:center;color:#ef4444;padding:20px">Error loading employees</td></tr>';
  }
}

function filterTable(){
  var srch = (document.getElementById('srch').value || '').toLowerCase();
  var mf   = document.getElementById('modeFilter').value;
  var of   = document.getElementById('officeFilter').value;
  var list = allEmps.filter(function(e){
    if(srch && !(e.name.toLowerCase().includes(srch) || e.emp_id.toLowerCase().includes(srch))) return false;
    if(mf && (e.location_mode||'office') !== mf) return false;
    if(of && e.office_setting !== of) return false;
    return true;
  });
  renderRows(list);
  document.getElementById('showCount').textContent = list.length + ' employees';
}

function renderRows(list){
  var tbody = document.getElementById('empBody');
  if(!list.length){
    tbody.innerHTML = '<tr><td colspan="9" style="text-align:center;padding:20px;color:var(--muted)">Koi employee nahi mila</td></tr>';
    return;
  }
  tbody.innerHTML = list.map(function(e){
    var mode = e.location_mode || 'office';
    var isHome = mode === 'home';
    var dis = isHome ? '' : 'disabled style="opacity:.4"';
    return '<tr data-emp="'+e.emp_id+'">' +
      '<td><input type="checkbox" class="emp-chk" value="'+e.emp_id+'" onchange="updateSelCount()" style="width:16px;height:16px;cursor:pointer"></td>' +
      '<td style="color:var(--muted);font-size:12px">'+e.emp_id+'</td>' +
      '<td><b>'+e.name+'</b></td>' +
      '<td style="font-size:12px;color:var(--muted)">'+(e.office_setting||'—')+'</td>' +
      '<td><select class="tbl-select" id="mode_'+e.emp_id+'" onchange="onModeChange(''+e.emp_id+'')">' +
        '<option value="office"'+(mode==='office'?' selected':'')+'>🏢 Office</option>' +
        '<option value="home"'+(mode==='home'?' selected':'')+'>🏠 Home/WFH</option>' +
        '<option value="any"'+(mode==='any'?' selected':'')+'>🌍 Field/Any</option>' +
      '</select></td>' +
      '<td><input type="number" step="0.000001" class="tbl-input" id="hlat_'+e.emp_id+'" value="'+(e.home_latitude||'')+'" placeholder="28.6139" '+(isHome?'':'disabled style="opacity:.4"')+'></td>' +
      '<td><input type="number" step="0.000001" class="tbl-input" id="hlng_'+e.emp_id+'" value="'+(e.home_longitude||'')+'" placeholder="77.2090" '+(isHome?'':'disabled style="opacity:.4"')+'></td>' +
      '<td><input type="number" class="tbl-input" id="hrad_'+e.emp_id+'" value="'+(e.home_radius||200)+'" '+(isHome?'':'disabled style="opacity:.4"')+'></td>' +
      '<td><button class="btn-save" onclick="saveEmp(''+e.emp_id+'')">💾</button></td>' +
    '</tr>';
  }).join('');
}

function onModeChange(eid){
  var mode = document.getElementById('mode_'+eid).value;
  var isHome = mode === 'home';
  ['hlat_','hlng_','hrad_'].forEach(function(p){
    var el = document.getElementById(p+eid);
    if(!el) return;
    el.disabled = !isHome;
    el.style.opacity = isHome ? '1' : '0.4';
  });
}

async function saveEmp(eid){
  var data = {
    emp_id:         eid,
    location_mode:  document.getElementById('mode_'+eid).value,
    home_latitude:  parseFloat(document.getElementById('hlat_'+eid).value) || null,
    home_longitude: parseFloat(document.getElementById('hlng_'+eid).value) || null,
    home_radius:    parseInt(document.getElementById('hrad_'+eid).value)   || 200
  };
  var res = await fetch('/api/employees/location/save', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(data)});
  var d   = await res.json();
  if(d.success){
    showToast('✅ Saved: ' + eid);
    var emp = allEmps.find(function(e){ return e.emp_id === eid; });
    if(emp){ emp.location_mode=data.location_mode; emp.home_latitude=data.home_latitude; emp.home_longitude=data.home_longitude; emp.home_radius=data.home_radius; }
  } else {
    showToast('❌ Error: ' + (d.error||''), false);
  }
}

function toggleAll(cb){
  document.querySelectorAll('.emp-chk').forEach(function(c){ c.checked = cb.checked; });
  updateSelCount();
}

function updateSelCount(){
  var n = document.querySelectorAll('.emp-chk:checked').length;
  document.getElementById('selCount').textContent = n + ' selected';
}

async function bulkApply(){
  var ids = Array.from(document.querySelectorAll('.emp-chk:checked')).map(function(c){ return c.value; });
  if(!ids.length){ showToast('Koi employee select nahi kiya!', false); return; }
  var mode = document.getElementById('bulkMode').value;
  var res = await fetch('/api/employees/location/bulk-save', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({emp_ids:ids, location_mode:mode})});
  var d   = await res.json();
  if(d.success){ showToast('✅ ' + d.updated + ' employees updated!'); loadEmps(); }
  else showToast('❌ Error', false);
}

loadEmps();
</script>
</body>
</html>
"""




# ===== EMPLOYEE LOCATION MANAGEMENT =====
@app.route('/api/employees/location/list', methods=['GET'])
@login_required
def api_emp_location_list():
    conn = get_db()
    emps = rows(conn.execute("""
        SELECT emp_id, name, department, office_setting,
               location_mode, home_latitude, home_longitude, home_radius
        FROM employees WHERE status='active' ORDER BY name
    """).fetchall())
    conn.close()
    return jsonify(emps)

@app.route('/api/employees/location/save', methods=['POST'])
@admin_required
def api_emp_location_save():
    data = request.json
    emp_id = data.get('emp_id')
    mode   = data.get('location_mode', 'office')
    home_lat = data.get('home_latitude')
    home_lng = data.get('home_longitude')
    home_rad = data.get('home_radius', 200)
    conn = get_db()
    conn.execute(
        'UPDATE employees SET location_mode=?, home_latitude=?, home_longitude=?, home_radius=? WHERE emp_id=?',
        (mode, home_lat, home_lng, home_rad, emp_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/employees/location/bulk-save', methods=['POST'])
@admin_required
def api_emp_location_bulk_save():
    data    = request.json
    emp_ids = data.get('emp_ids', [])
    mode    = data.get('location_mode', 'office')
    conn    = get_db()
    for eid in emp_ids:
        conn.execute('UPDATE employees SET location_mode=? WHERE emp_id=?', (mode, eid))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'updated': len(emp_ids)})

# ===== MOBILE / PWA =====
@app.route('/my-attendance')
@login_required
def my_attendance():
    conn    = get_db()
    u       = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    emp_id  = u['emp_id'] if u else None
    emp     = None
    records = []
    if emp_id:
        emp = conn.execute('SELECT * FROM employees WHERE emp_id=?', (emp_id,)).fetchone()
        if emp: emp = dict(emp)
        # Last 30 days attendance
        records = rows(conn.execute(
            'SELECT * FROM attendance WHERE emp_id=? ORDER BY date DESC LIMIT 60', (emp_id,)
        ).fetchall())
    conn.close()
    today = date.today().isoformat()
    return render_template('my_attendance.html', emp=emp, records=records, today=today, emp_id=emp_id)

@app.route('/api/my/mark', methods=['POST'])
@login_required
def api_my_mark():
    """User apni khud ki attendance mark kare with time + location"""
    import math
    from datetime import datetime as dt
    u = session.get('emp_id') or ''
    if not u:
        conn2 = get_db()
        row2  = conn2.execute('SELECT emp_id FROM users WHERE id=?', (session['user_id'],)).fetchone()
        conn2.close()
        u = row2['emp_id'] if row2 else None
    if not u:
        return jsonify({'error': 'Employee ID linked nahi hai. Admin se contact karo.'}), 400

    data    = request.json
    today   = date.today().isoformat()
    now_time= dt.now().strftime('%H:%M:%S')
    status  = data.get('status', 'P')
    remark  = data.get('remark', '').strip()
    lat     = data.get('latitude')
    lng     = data.get('longitude')
    address = data.get('address', '').strip()

    # Check location against per-employee mode
    location_status = 'not_checked'

    def haversine(lat1, lon1, lat2, lon2):
        R = 6371000
        phi1, phi2 = math.radians(lat1), math.radians(lat2)
        dphi = math.radians(lat2 - lat1)
        dlambda = math.radians(lon2 - lon1)
        a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlambda/2)**2
        return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))

    if lat and lng:
        conn_loc = get_db()
        emp_rec = row(conn_loc.execute('SELECT location_mode, home_latitude, home_longitude, home_radius FROM employees WHERE emp_id=?', (u,)).fetchone())
        loc_settings = rows(conn_loc.execute('SELECT * FROM location_settings WHERE is_active=1').fetchall())
        conn_loc.close()

        mode = (emp_rec.get('location_mode') or 'office') if emp_rec else 'office'

        if mode == 'any':
            # Field employee — no restriction
            location_status = 'in_range'
        elif mode == 'home' and emp_rec and emp_rec.get('home_latitude') and emp_rec.get('home_longitude'):
            # WFH — check against home coords
            dist = haversine(float(lat), float(lng), float(emp_rec['home_latitude']), float(emp_rec['home_longitude']))
            radius = float(emp_rec.get('home_radius') or 200)
            location_status = 'in_range' if dist <= radius else 'out_of_range'
        elif mode == 'home' and emp_rec and not emp_rec.get('home_latitude'):
            # WFH but home not set yet
            location_status = 'home_not_set'
        else:
            # Office mode — check against office location settings
            if loc_settings:
                location_status = 'out_of_range'
                for loc in loc_settings:
                    dist = haversine(float(lat), float(lng), float(loc['latitude']), float(loc['longitude']))
                    if dist <= float(loc['radius_meters']):
                        location_status = 'in_range'
                        break
            else:
                location_status = 'not_checked'

    conn = get_db()
    existing = conn.execute('SELECT id FROM attendance WHERE emp_id=? AND date=?', (u, today)).fetchone()
    if existing:
        conn.execute(
            'UPDATE attendance SET status=?,remark=?,marked_by=?,marked_at=CURRENT_TIMESTAMP,check_in_time=?,latitude=?,longitude=?,location_address=?,location_status=? WHERE id=?',
            (status, remark, session['username'], now_time, lat, lng, address, location_status, existing['id'])
        )
    else:
        conn.execute(
            'INSERT INTO attendance (emp_id,date,status,remark,marked_by,check_in_time,latitude,longitude,location_address,location_status) VALUES (?,?,?,?,?,?,?,?,?,?)',
            (u, today, status, remark, session['username'], now_time, lat, lng, address, location_status)
        )
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'date': today, 'status': status, 'time': now_time, 'location_status': location_status})

@app.route('/profile')
@login_required
def profile():
    conn = get_db()
    u = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    if not u:
        conn.close()
        return redirect(url_for('login'))
    u = dict(u)
    emp = None
    if u.get('emp_id'):
        emp = conn.execute('SELECT * FROM employees WHERE emp_id=?', (u['emp_id'],)).fetchone()
        if emp: emp = dict(emp)
    conn.close()
    return render_template('profile.html', user=u, emp=emp)

@app.route('/mobile')
def mobile():
    return render_template('mobile_attendance.html')

@app.route('/api/me')
def api_me():
    if 'user_id' not in session:
        return jsonify({'error': 'not logged in'}), 401
    conn = get_db()
    emp = conn.execute('SELECT emp_id FROM employees WHERE emp_id=? OR email=? LIMIT 1',
                       (session.get('username',''), session.get('username',''))).fetchone()
    conn.close()
    return jsonify({
        'user_id':  session['user_id'],
        'username': session['username'],
        'name':     session.get('name', ''),
        'role':     session.get('role', 'viewer'),
        'emp_id':   emp['emp_id'] if emp else session['username']
    })

@app.route('/mobile-setup')
@login_required
def mobile_setup():
    if session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    return render_template('mobile_setup.html')

@app.route('/manifest.json')
def manifest():
    return send_file('static/manifest.json', mimetype='application/manifest+json')

@app.route('/api/roster/apply-month', methods=['POST'])
@login_required
def api_roster_apply_month():
    """Apply employee templates to entire month with smart Saturday rule"""
    import calendar
    data         = request.json
    year         = int(data.get('year', 0))
    month        = int(data.get('month', 0))
    sat_rule     = data.get('sat_rule', 'standard')  # standard|all_on|all_off|alternate
    overwrite    = data.get('overwrite', True)
    emp_ids      = data.get('emp_ids', [])  # empty = all

    if not year or not month:
        return jsonify({'success': False, 'error': 'year and month required'})

    conn = get_db()

    # Get all employee rules (with template data)
    if emp_ids:
        ph    = ','.join(['?']*len(emp_ids))
        rules = rows(conn.execute(f'SELECT * FROM roster_rules WHERE emp_id IN ({ph})', emp_ids).fetchall())
    else:
        rules = rows(conn.execute('SELECT * FROM roster_rules').fetchall())

    if not rules:
        conn.close()
        return jsonify({'success': False, 'error': 'No employee mappings found'})

    # Find all Saturdays in the month and their occurrence number
    first_day, days_in_month = calendar.monthrange(year, month)
    saturdays = []
    for day in range(1, days_in_month + 1):
        if calendar.date(year, month, day).weekday() == 5:  # 5 = Saturday
            saturdays.append(day)

    # Build Saturday rule: which occurrence numbers are ON
    def sat_shift(occ, total_sats, rule):
        if rule == 'all_on':   return 'W'
        if rule == 'all_off':  return 'WO'
        if rule == 'alternate': return 'W' if occ % 2 == 1 else 'WO'
        # standard: 1st ON, 2nd-4th OFF, 5th ON
        if rule == 'standard':
            if occ == 1: return 'W'
            if occ == total_sats and total_sats == 5: return 'W'
            return 'WO'
        return 'WO'

    # Map each Saturday date → its shift
    sat_shift_map = {}
    for idx, sat_day in enumerate(saturdays):
        sat_shift_map[sat_day] = sat_shift(idx+1, len(saturdays), sat_rule)

    # Get all Monday-starting weeks that overlap this month
    from datetime import date, timedelta
    def get_monday(d):
        return d - timedelta(days=d.weekday())

    weeks = set()
    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        weeks.add(get_monday(d))
    weeks = sorted(weeks)

    saved = skipped = 0
    for rule in rules:
        eid = rule['emp_id']
        for week_monday in weeks:
            week_start = str(week_monday)
            # Build 7 days for this week
            shifts = {}
            day_keys = ['mon','tue','wed','thu','fri','sat','sun']
            for i, dk in enumerate(day_keys):
                day_date = week_monday + timedelta(days=i)
                # Only apply if day is in the target month
                if day_date.month != month:
                    # Use template default for days outside month
                    shifts[dk] = rule.get(dk, 'W')
                    continue
                if dk == 'sat':
                    shifts[dk] = sat_shift_map.get(day_date.day, rule.get('sat','WO'))
                else:
                    shifts[dk] = rule.get(dk, 'W')

            existing = conn.execute('SELECT id FROM roster WHERE emp_id=? AND week_start=?', (eid, week_start)).fetchone()
            if existing and not overwrite:
                skipped += 1
                continue
            if existing:
                conn.execute(
                    'UPDATE roster SET mon=?,tue=?,wed=?,thu=?,fri=?,sat=?,sun=? WHERE emp_id=? AND week_start=?',
                    (shifts['mon'],shifts['tue'],shifts['wed'],shifts['thu'],shifts['fri'],shifts['sat'],shifts['sun'], eid, week_start)
                )
            else:
                conn.execute(
                    'INSERT INTO roster (emp_id,week_start,mon,tue,wed,thu,fri,sat,sun) VALUES (?,?,?,?,?,?,?,?,?)',
                    (eid, week_start, shifts['mon'],shifts['tue'],shifts['wed'],shifts['thu'],shifts['fri'],shifts['sat'],shifts['sun'])
                )
            saved += 1

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'saved': saved, 'skipped': skipped,
                    'weeks': len(weeks), 'employees': len(rules),
                    'saturdays': len(saturdays), 'sat_rule': sat_rule})

@app.route('/api/roster/apply-pattern', methods=['POST'])
@login_required
def api_roster_apply_pattern():
    """Apply monthly pattern: 1st & last Saturday = Working, rest = WO, Mon-Fri = W, Sun = WO"""
    from datetime import date, timedelta
    data    = request.json
    year    = int(data.get('year', date.today().year))
    month   = int(data.get('month', date.today().month))
    office  = data.get('office', '')        # optional filter
    emp_ids = data.get('emp_ids', [])       # optional specific employees
    overwrite = data.get('overwrite', False)

    conn = get_db()

    # Get employees
    if emp_ids:
        emps = [r['emp_id'] for r in rows(conn.execute(
            f"SELECT emp_id FROM employees WHERE status='active' AND emp_id IN ({','.join(['?']*len(emp_ids))})",
            emp_ids).fetchall())]
    elif office:
        emps = [r['emp_id'] for r in rows(conn.execute(
            "SELECT emp_id FROM employees WHERE status='active' AND office_setting=?", (office,)).fetchall())]
    else:
        emps = [r['emp_id'] for r in rows(conn.execute(
            "SELECT emp_id FROM employees WHERE status='active'").fetchall())]

    # Get all Saturdays in month
    def get_saturdays(y, m):
        sats = []
        d = date(y, m, 1)
        while d.weekday() != 5: d += timedelta(days=1)
        while d.month == m:
            sats.append(d); d += timedelta(days=7)
        return sats

    sats = get_saturdays(year, month)
    if not sats:
        conn.close()
        return jsonify({'success': False, 'error': 'No Saturdays found'})

    first_sat = sats[0]
    last_sat  = sats[-1]

    # Get all weeks in month
    def get_weeks(y, m):
        weeks = []
        d = date(y, m, 1)
        day = d.weekday()
        d -= timedelta(days=day)  # Go to Monday
        last_day = date(y, m, [31,28+((y%4==0 and y%100!=0) or y%400==0),31,30,31,30,31,31,30,31,30,31][m-1], )
        while d <= last_day:
            weeks.append(d)
            d += timedelta(days=7)
        return weeks

    weeks = get_weeks(year, month)
    saved = skipped = 0

    for emp_id in emps:
        for week_mon in weeks:
            week_start = week_mon.strftime('%Y-%m-%d')
            sat_date   = week_mon + timedelta(days=5)

            # Determine SAT shift: 1st or last Saturday of the month = W, else WO
            sat_shift = 'W' if (sat_date == first_sat or sat_date == last_sat) else 'WO'

            # Check if record exists
            existing = conn.execute(
                'SELECT id FROM roster WHERE emp_id=? AND week_start=?',
                (emp_id, week_start)).fetchone()

            if existing and not overwrite:
                skipped += 1
                continue

            shifts = {'mon':'W','tue':'W','wed':'W','thu':'W','fri':'W','sat':sat_shift,'sun':'WO'}

            if existing:
                conn.execute(
                    'UPDATE roster SET mon=?,tue=?,wed=?,thu=?,fri=?,sat=?,sun=? WHERE emp_id=? AND week_start=?',
                    (shifts['mon'],shifts['tue'],shifts['wed'],shifts['thu'],shifts['fri'],
                     shifts['sat'],shifts['sun'], emp_id, week_start))
            else:
                conn.execute(
                    'INSERT INTO roster (emp_id,week_start,mon,tue,wed,thu,fri,sat,sun) VALUES (?,?,?,?,?,?,?,?,?)',
                    (emp_id,week_start,shifts['mon'],shifts['tue'],shifts['wed'],shifts['thu'],
                     shifts['fri'],shifts['sat'],shifts['sun']))
            saved += 1

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'saved': saved, 'skipped': skipped,
                    'first_sat': first_sat.strftime('%d %b'), 'last_sat': last_sat.strftime('%d %b')})


if __name__ == '__main__':
    os.makedirs('database', exist_ok=True)
    init_db()
    print("\n" + "="*50)
    print("  EasySource HRMS Enterprise Server")
    print("  URL: http://localhost:5000")
    print("  Admin: admin / admin123")
    print("="*50 + "\n")
    app.run(host='0.0.0.0', port=5000, debug=True)
